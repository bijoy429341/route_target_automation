import streamlit as st
import pandas as pd
import numpy as np
import hashlib
from io import BytesIO
import os
import sys
import time
import zipfile
import re
import tempfile
import difflib
from datetime import datetime
from typing import Union
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

# ======================= PAGE CONFIG =======================
st.set_page_config(
    page_title="Route Wise Target Automation",
    layout="wide",
    page_icon="🎯",
    initial_sidebar_state="expanded"
)

# ======================= GLOBAL CSS =======================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
* { font-family: 'Inter', sans-serif; }
.stApp {
    background: linear-gradient(135deg, #0a0e1a 0%, #111827 60%, #0a0e1a 100%);
    color: #e2e8f0;
}
.main .block-container {
    background: rgba(17,24,39,0.97);
    border-radius: 18px;
    padding: 2rem 2.5rem;
    margin-top: 0.5rem;
    box-shadow: 0 20px 60px rgba(0,0,0,0.6), inset 0 1px 0 rgba(255,255,255,0.05);
    border: 1px solid rgba(99,179,237,0.08);
}
.hero-banner {
    background: linear-gradient(135deg, #1a365d 0%, #2b6cb0 50%, #1a365d 100%);
    border: 1px solid #3182ce;
    border-radius: 16px;
    padding: 1.4rem 2rem;
    margin-bottom: 1.5rem;
    display: flex;
    align-items: center;
    gap: 1rem;
    box-shadow: 0 8px 32px rgba(49,130,206,0.25);
}
.hero-title { font-size: 1.6rem; font-weight: 700; color: #ebf8ff; margin: 0; }
.hero-sub { font-size: 0.82rem; color: #bee3f8; margin: 0.3rem 0 0 0; }
.hero-badges span {
    background: rgba(99,179,237,0.2); border: 1px solid rgba(99,179,237,0.4);
    border-radius: 20px; padding: 2px 10px; font-size: 0.72rem;
    color: #90cdf4; margin-right: 6px;
}
.stTabs [data-baseweb="tab-list"] {
    background: rgba(26,32,44,0.9); border-radius: 14px;
    padding: 6px; border: 1px solid rgba(99,179,237,0.12); gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    background: transparent; border-radius: 10px; color: #a0aec0;
    font-weight: 500; font-size: 0.82rem; padding: 8px 14px;
    border: 1px solid transparent; transition: all 0.25s ease;
}
.stTabs [data-baseweb="tab"]:hover {
    background: rgba(49,130,206,0.12); color: #90cdf4;
    border-color: rgba(49,130,206,0.2);
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #2b6cb0 0%, #1a4c8b 100%) !important;
    color: #fff !important; border-color: #3182ce !important;
    box-shadow: 0 4px 14px rgba(49,130,206,0.4);
}
.stButton > button {
    background: linear-gradient(135deg, #2b6cb0 0%, #1a4c8b 100%);
    color: white; border: none; border-radius: 10px; font-weight: 600;
    font-size: 0.85rem; padding: 0.55rem 1.2rem;
    transition: all 0.25s ease; box-shadow: 0 4px 14px rgba(49,130,206,0.3);
}
.stButton > button:hover {
    background: linear-gradient(135deg, #3182ce 0%, #2b6cb0 100%);
    transform: translateY(-2px); box-shadow: 0 8px 24px rgba(49,130,206,0.45);
}
.stButton > button[kind="secondary"] {
    background: rgba(45,55,72,0.9); border: 1px solid rgba(99,179,237,0.3); box-shadow: none;
}
/* ── Per-step uploader card ─────────────────── */
.prev-step-card {
    background: linear-gradient(135deg, rgba(26,32,44,0.98) 0%, rgba(17,26,38,0.98) 100%);
    border: 1.5px solid rgba(99,179,237,0.22);
    border-radius: 14px;
    padding: 1.1rem 1.4rem 1rem 1.4rem;
    margin-bottom: 1.1rem;
    position: relative;
}
.prev-step-card-title {
    font-size: 0.78rem; font-weight: 700; letter-spacing: 0.08em;
    text-transform: uppercase; color: #63b3ed; margin-bottom: 0.3rem;
}
.prev-step-card-desc {
    font-size: 0.78rem; color: #718096; margin-bottom: 0.55rem;
}
.prev-step-badge {
    display: inline-block;
    background: rgba(49,130,206,0.15); border: 1px solid rgba(49,130,206,0.35);
    border-radius: 20px; padding: 2px 10px; font-size: 0.7rem; color: #90cdf4;
    margin-right: 5px; margin-bottom: 4px;
}
.prev-step-badge.green {
    background: rgba(72,187,120,0.15); border-color: rgba(72,187,120,0.4); color: #68d391;
}
.session-banner {
    background: linear-gradient(90deg, rgba(72,187,120,0.10) 0%, rgba(26,32,44,0.85) 100%);
    border: 1px solid rgba(72,187,120,0.3); border-radius: 10px;
    padding: 0.6rem 1rem; margin: 0.4rem 0 0.7rem 0;
    font-size: 0.82rem; color: #68d391;
}
.override-hint {
    font-size: 0.73rem; color: #4a5568; margin-top: 0.25rem;
}
/* ───────────────────────────────────────────── */
.info-card {
    background: linear-gradient(135deg, rgba(43,108,176,0.1) 0%, rgba(26,32,44,0.95) 100%);
    border: 1px solid rgba(99,179,237,0.2); border-radius: 14px;
    padding: 1.2rem 1.5rem; margin: 0.6rem 0;
}
.success-card {
    background: linear-gradient(135deg, rgba(72,187,120,0.1) 0%, rgba(26,32,44,0.95) 100%);
    border: 1px solid rgba(72,187,120,0.3); border-radius: 14px;
    padding: 1rem 1.4rem; margin: 0.6rem 0;
}
.upload-section {
    background: rgba(26,32,44,0.5);
    border: 1px solid rgba(99,179,237,0.12);
    border-radius: 14px; padding: 1.2rem 1.5rem; margin-bottom: 1rem;
}
.step-header {
    background: linear-gradient(90deg, rgba(43,108,176,0.15) 0%, transparent 100%);
    border-left: 4px solid #3182ce; border-radius: 0 10px 10px 0;
    padding: 0.8rem 1.2rem; margin-bottom: 1.2rem;
}
.step-header h3 { color: #90cdf4; margin: 0; font-size: 1.15rem; }
.step-header p { color: #718096; margin: 0.3rem 0 0 0; font-size: 0.8rem; }
div[data-testid="metric-container"] {
    background: rgba(26,32,44,0.9); border: 1px solid rgba(99,179,237,0.15);
    border-radius: 12px; padding: 0.9rem 1rem; box-shadow: 0 4px 12px rgba(0,0,0,0.2);
}
div[data-testid="metric-container"] label { color: #718096 !important; font-size: 0.75rem !important; }
div[data-testid="metric-container"] div[data-testid="stMetricValue"] { color: #90cdf4 !important; font-weight: 700; }
.stSelectbox > div > div,
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stTextArea > div > div > textarea {
    background: rgba(26,32,44,0.95) !important;
    border: 1px solid rgba(74,85,104,0.6) !important;
    border-radius: 8px !important; color: #e2e8f0 !important;
}
.stMultiSelect > div > div {
    background: rgba(26,32,44,0.95) !important;
    border: 1px solid rgba(74,85,104,0.6) !important; border-radius: 8px !important;
}
.streamlit-expanderHeader {
    background: rgba(26,32,44,0.9) !important;
    border: 1px solid rgba(74,85,104,0.4) !important;
    border-radius: 10px !important; color: #a0aec0 !important;
}
.stSuccess > div { background: rgba(72,187,120,0.12) !important; border-radius: 10px; border-left: 4px solid #48bb78; }
.stInfo > div { background: rgba(66,153,225,0.12) !important; border-radius: 10px; border-left: 4px solid #4299e1; }
.stWarning > div { background: rgba(237,137,54,0.12) !important; border-radius: 10px; border-left: 4px solid #ed8936; }
.stError > div { background: rgba(245,101,101,0.12) !important; border-radius: 10px; border-left: 4px solid #f56565; }
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0d1117 0%, #111827 100%) !important;
    border-right: 1px solid rgba(99,179,237,0.12) !important;
}
section[data-testid="stSidebar"] .stButton > button { width: 100%; font-size: 0.8rem; }
.sidebar-step {
    background: rgba(26,32,44,0.7); border: 1px solid rgba(74,85,104,0.3);
    border-radius: 10px; padding: 0.55rem 0.9rem; margin: 0.25rem 0; font-size: 0.82rem;
}
.sidebar-step.done { border-color: rgba(72,187,120,0.4); background: rgba(72,187,120,0.08); }
.stDataFrame > div { border-radius: 12px; overflow: hidden; border: 1px solid rgba(74,85,104,0.3); }
hr { border-color: rgba(74,85,104,0.3) !important; }
.stFileUploader > div {
    background: rgba(26,32,44,0.7) !important;
    border: 2px dashed rgba(99,179,237,0.3) !important; border-radius: 12px !important;
}
.stProgress > div > div { background: linear-gradient(90deg,#2b6cb0,#63b3ed) !important; border-radius: 4px; }
::-webkit-scrollbar { width:6px; height:6px; }
::-webkit-scrollbar-track { background: rgba(26,32,44,0.5); border-radius: 4px; }
::-webkit-scrollbar-thumb { background: #4a5568; border-radius: 4px; }
.chip {
    display: inline-block; background: rgba(49,130,206,0.15);
    border: 1px solid rgba(49,130,206,0.35); border-radius: 20px;
    padding: 2px 10px; font-size: 0.72rem; color: #90cdf4; margin: 2px;
}
.chip.green { background:rgba(72,187,120,0.15); border-color:rgba(72,187,120,0.4); color:#68d391; }
</style>
""", unsafe_allow_html=True)

# ── Hero ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero-banner">
  🎯
  <div>
    <h1 class="hero-title">Route Wise Target Automation</h1>
    <p class="hero-sub">Template → AVG → Merge → Split → Area → Email  |  Every tab works independently — upload the previous step result to begin</p>
    <div class="hero-badges">
      <span>⚡ Fast</span><span>🔁 Session Carry-forward</span><span>📊 VBA-exact Logic</span><span>📍 xlsx · xlsb · csv Area Split</span><span>📧 Outlook</span>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# =============================================================================
# SHARED UTILITIES
# =============================================================================
@st.cache_data(show_spinner=False, max_entries=128)
def _read_tabular_bytes(file_name: str, blob: bytes, is_csv: bool) -> pd.DataFrame:
    bio = BytesIO(blob)
    if is_csv:
        return pd.read_csv(bio, low_memory=False)
    if file_name.lower().endswith(".xlsb"):
        return pd.read_excel(bio, engine="pyxlsb")
    return pd.read_excel(bio, engine="openpyxl")

def load_file(file):
    if file is None:
        return None
    try:
        blob = file.getvalue()
        name = str(getattr(file, "name", "") or "upload")
        is_csv = name.lower().endswith(".csv")
        return _read_tabular_bytes(name, blob, is_csv)
    except Exception as e:
        raise RuntimeError(f"File read failed ({getattr(file,'name','?')}): {e}") from e

def clean_filename(name):
    return re.sub(r'[\\/:*?"<>|]', "_", str(name))

def _normalize_header(t):
    return str(t).lower().replace(" ", "").replace("_", "").replace("-", "")

def _guess(cols, keywords, fallback=0):
    for i, c in enumerate(cols):
        if any(k in str(c).lower() for k in keywords):
            return i
    return min(fallback, max(len(cols) - 1, 0))

def get_column_suggestions(df, column_type="general"):
    if df is None or df.empty:
        return []
    mapping = {
        "location": ["route","location","territory","area","zone","db loc"],
        "product":  ["product","prod","item","sku","code"],
        "month":    ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"],
        "quantity": ["qty","quantity","amount","value","sales"],
        "target":   ["target","db","distributor"],
        "segment":  ["segment","seg","category","type"],
    }
    kws = mapping.get(column_type, [])
    return [c for c in df.columns if any(k in str(c).lower() for k in kws)] if kws else df.columns.tolist()

# ── Session helpers ───────────────────────────────────────────────────────────
def ss_set(k, v): st.session_state[k] = v
def ss_get(k, d=None): return st.session_state.get(k, d)
def mark_step_done(n): done = ss_get("steps_done", set()); done.add(n); ss_set("steps_done", done)
def step_done(n): return n in ss_get("steps_done", set())

# ── UI helpers ────────────────────────────────────────────────────────────────
def step_header(icon, title, caption=""):
    st.markdown(
        f'<div class="step-header"><h3>{icon} {title}</h3>'
        + (f'<p>{caption}</p>' if caption else '') + '</div>',
        unsafe_allow_html=True)

def mini_dashboard(df, title="📊 Overview", key="d"):
    if df is None or df.empty:
        st.info("No data."); return
    st.markdown(f"**{title}**")
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Rows", f"{len(df):,}")
    c2.metric("Columns", f"{len(df.columns)}")
    c3.metric("Numeric", f"{len(df.select_dtypes(include=[np.number]).columns)}")
    miss = (df.isnull().sum().sum() / max(len(df)*len(df.columns),1))*100
    c4.metric("Missing%", f"{miss:.1f}%")
    sel = st.selectbox("Inspect column", ["— pick —"]+df.columns.tolist(), key=f"dc_{key}")
    if sel != "— pick —":
        cd = df[sel]
        cx1,cx2 = st.columns(2)
        with cx1:
            st.write(f"Type `{cd.dtype}` · Unique `{cd.nunique():,}` · Missing `{cd.isnull().sum():,}`")
        with cx2:
            if pd.api.types.is_numeric_dtype(cd):
                st.write(f"Min `{cd.min():,.2f}` · Max `{cd.max():,.2f}` · Mean `{cd.mean():,.2f}`")
            else:
                for v, n in cd.value_counts().head(5).items():
                    st.write(f"`{v}` — {n:,}")

def download_csv(df, filename, label="⬇️ Download CSV", key=None):
    st.download_button(label, df.to_csv(index=False).encode("utf-8"), filename, "text/csv", key=key)

def download_excel(df, filename, label="📊 Download Excel", key=None):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False)
    st.download_button(label, buf.getvalue(), filename,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=key)

def required_error(msg):
    st.error(f"⚠️ Required: {msg}")


# =============================================================================
# ██████████████████████████████████████████████████████████████████████████████
#  INDIVIDUAL PREVIOUS-STEP UPLOADER
#  Each tab calls this once at the top with its own config.
#  Returns the resolved DataFrame (from upload OR session) or None.
# ██████████████████████████████████████████████████████████████████████████████
def prev_step_uploader(
    *,
    tab_key: str,           # unique key prefix  e.g. "t2"
    from_step: int,         # step number being consumed  e.g. 1
    from_label: str,        # human label  e.g. "Step 1 — Template"
    from_desc: str,         # one-line description of what the file contains
    sess_key: str,          # session_state key where the df lives
    accepted_cols_hint: str = "",   # short hint about required columns (optional)
    file_types: list = None,
) -> Union[pd.DataFrame, None]:
    """
    Renders a self-contained 'Previous Step Result' uploader card.
    • If session already holds data  → shows a green 'loaded from session' banner
    • User can always upload a fresh file to override
    Returns the active DataFrame or None.
    """
    file_types = file_types or ["csv", "xlsx", "xlsb"]
    sess_df: Union[pd.DataFrame, None] = ss_get(sess_key)

    st.markdown(
        f"""
        <div class="prev-step-card">
          <div class="prev-step-card-title">
            📥 Previous Step Input &nbsp;·&nbsp; Step {from_step} Result
          </div>
          <div class="prev-step-card-desc">
            <strong>{from_label}</strong> — {from_desc}
            {"<br><span style='color:#4a5568;font-size:0.72rem;'>Expected columns: "+accepted_cols_hint+"</span>" if accepted_cols_hint else ""}
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Session banner ────────────────────────────────────────────────────────
    if sess_df is not None:
        rows = f"{len(sess_df):,} rows × {len(sess_df.columns)} cols"
        st.markdown(
            f'<div class="session-banner">✅ Loaded from session — {rows}'
            f'<span class="override-hint"> · Upload below to override</span></div>',
            unsafe_allow_html=True,
        )

    # ── File uploader ─────────────────────────────────────────────────────────
    uploaded = st.file_uploader(
        f"Upload {from_label} (CSV / Excel)",
        type=file_types,
        key=f"{tab_key}_prev_upload",
        help=f"Re-upload the output of Step {from_step} to override session data.",
    )

    if uploaded is not None:
        try:
            df_new = load_file(uploaded)
            ss_set(sess_key, df_new)
            mark_step_done(from_step)
            rows = f"{len(df_new):,} rows × {len(df_new.columns)} cols"
            st.success(f"✅ **{uploaded.name}** loaded — {rows}")
            with st.expander("📊 Quick preview", expanded=False):
                st.dataframe(df_new.head(10), use_container_width=True)
            return df_new
        except RuntimeError as e:
            st.error(str(e))
            return sess_df   # fall back to session if upload fails

    return sess_df   # None if no session and no upload


# =============================================================================
# CACHED LOADERS
# =============================================================================
@st.cache_data(show_spinner=False)
def _load_t2(name: str, blob: bytes) -> pd.DataFrame:
    bio = BytesIO(blob)
    if name.lower().endswith(".csv"):
        return pd.read_csv(bio, low_memory=False)
    return pd.read_excel(bio, engine="openpyxl")

@st.cache_data(show_spinner=False)
def _load_fmt(name: str, blob: bytes) -> pd.DataFrame:
    bio = BytesIO(blob)
    if name.lower().endswith(".csv"):
        return pd.read_csv(bio, dtype=str)
    return pd.read_excel(bio, dtype=str)

# =============================================================================
# STEP 4 — Route split (VBA-exact)
# =============================================================================
def vba_val(value) -> float:
    """Mimic VBA Val(): read the leading numeric part, otherwise return 0."""
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float, np.integer, np.floating)):
        return 0.0 if pd.isna(value) else float(value)
    text = str(value).lstrip()
    match = re.match(r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?", text)
    return float(match.group(0)) if match else 0.0

def vba_cell_text(value) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)):
        return str(int(value)) if float(value).is_integer() else str(float(value))
    return str(value)

def vba_is_numeric(value) -> bool:
    if value is None or pd.isna(value):
        return False
    if isinstance(value, (int, float, np.integer, np.floating)):
        return True
    return re.match(r"^\s*[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?\s*$", str(value)) is not None

def run_route_split(df, loc_col, prod_col, route_seg_col, prod_seg_col,
                    contri_col, target_cols, radhuni_strict, ruchi_strict):
    start = time.time()
    result_df = df.copy()
    total_source = total_split = 0.0
    gap_dict = {}
    row_indices = list(result_df.index)

    for tgt_col in target_cols:
        new_col = f"Route Wise {tgt_col}"
        result_df[new_col] = 0.0

        groups = {}
        for row_idx in row_indices:
            loc_text = vba_cell_text(result_df.at[row_idx, loc_col])
            prod_text = vba_cell_text(result_df.at[row_idx, prod_col])
            grp_key = f"{loc_text}|{prod_text}"
            groups.setdefault(grp_key, []).append(row_idx)

        for grp_key, group_rows in groups.items():
            first_row = group_rows[0]
            total_qty = vba_val(result_df.at[first_row, tgt_col])
            total_source += total_qty

            adjustable_rows = []
            total_contri = 0.0
            for row_idx in group_rows:
                route_seg = vba_cell_text(result_df.at[row_idx, route_seg_col]).strip().lower()
                prod_seg = vba_cell_text(result_df.at[row_idx, prod_seg_col]).strip().lower()

                if (
                    route_seg == "combine"
                    or (route_seg == "radhuni" and ((not radhuni_strict) or route_seg == prod_seg))
                    or (route_seg == "ruchi" and ((not ruchi_strict) or route_seg == prod_seg))
                    or route_seg == prod_seg
                ):
                    adjustable_rows.append(row_idx)
                    total_contri += vba_val(result_df.at[row_idx, contri_col])

            for row_idx in group_rows:
                result_df.at[row_idx, new_col] = 0

            if len(adjustable_rows) == 0:
                gkey = (
                    f"DB LOCATION CODE:{vba_cell_text(result_df.at[first_row, loc_col])}"
                    f" | SKU:{vba_cell_text(result_df.at[first_row, prod_col])}"
                )
                gap_dict[gkey] = gap_dict.get(gkey, 0.0) + total_qty
                continue

            n = len(adjustable_rows)
            if total_contri <= 0:
                base = int(total_qty / n)
                remaining = int(total_qty - (base * n))
                split_vals = np.full(n, base, dtype=float)
                for i in range(max(remaining, 0)):
                    if i < n:
                        split_vals[i] += 1
            else:
                split_vals = np.zeros(n, dtype=float)
                frac = np.zeros(n, dtype=float)
                remaining = total_qty
                for adj_pos, row_idx in enumerate(adjustable_rows):
                    contrib_val = vba_val(result_df.at[row_idx, contri_col])
                    raw_qty = total_qty * contrib_val / total_contri
                    split_vals[adj_pos] = int(raw_qty)
                    remaining -= split_vals[adj_pos]
                remaining = int(remaining)
                for adj_pos, row_idx in enumerate(adjustable_rows):
                    contrib_val = vba_val(result_df.at[row_idx, contri_col])
                    raw_qty = total_qty * contrib_val / total_contri
                    frac[adj_pos] = raw_qty - int(raw_qty)
                for _ in range(max(remaining, 0)):
                    max_idx = 0
                    max_frac = frac[0]
                    for adj_pos in range(1, n):
                        if frac[adj_pos] > max_frac:
                            max_frac = frac[adj_pos]
                            max_idx = adj_pos
                    split_vals[max_idx] += 1
                    frac[max_idx] = 0

            for adj_pos, row_idx in enumerate(adjustable_rows):
                result_df.at[row_idx, new_col] = split_vals[adj_pos]
                total_split += split_vals[adj_pos]

    route_cols = [f"Route Wise {t}" for t in target_cols]
    result_df["Total Route Wise Target Qty"] = result_df[route_cols].sum(axis=1)
    return result_df, total_source, total_split, gap_dict, time.time() - start

# =============================================================================
# STEP 5 — Area split helpers
# =============================================================================
def normalize_output_folder(path):
    if path is None: return ""
    s = str(path).strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ('"', "'"): s = s[1:-1].strip()
    return s

def is_area_code_col(c): n = _normalize_header(c); return "area" in n and "code" in n
def is_area_name_col(c): n = _normalize_header(c); return "area" in n and "name" in n
def is_unit_column(c): return any(k in _normalize_header(c) for k in ["unit","uinit","unt"])
def is_value_column(c): return any(k in _normalize_header(c) for k in ["value","vlue","val","amount","tk","taka"])
def is_market_column(c): return any(k in _normalize_header(c) for k in ["market","mkt","thana","outlet"])
def is_territory_column(c): return any(k in _normalize_header(c) for k in ["territory","terri","tso","t.s.o"])
def is_product_column(c): return any(k in _normalize_header(c) for k in ["product","prod","item","sku"])
def is_category_column(c): return any(k in _normalize_header(c) for k in ["category","cat","segment"])
def is_pack_size_column(c): return any(k in _normalize_header(c) for k in ["packsize","pack","size","gram","gm","kg","ml","ltr","liter"])

def _numeric_series(df, col):
    cleaned = df[col].astype(str).str.replace(",", "", regex=False).str.replace("%", "", regex=False).str.strip()
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)

def find_target_value_column(df):
    cols = df.columns.tolist()
    numeric_cols = [
        c for c in cols
        if pd.api.types.is_numeric_dtype(df[c]) or _numeric_series(df, c).abs().sum() > 0
    ]
    if not numeric_cols:
        return None

    def score(col):
        n = _normalize_header(col)
        s = 0
        if "targetvalue" in n: s += 120
        if "target" in n: s += 80
        if any(k in n for k in ["value", "vlue", "val", "amount", "taka", "tk"]): s += 50
        if n.startswith("routewise"): s += 20
        if any(k in n for k in ["qty", "quantity", "unit", "uinit", "unt", "code", "id"]): s -= 40
        return s

    ranked = sorted(numeric_cols, key=lambda c: (score(c), _numeric_series(df, c).abs().sum()), reverse=True)
    return ranked[0] if score(ranked[0]) > 0 else None

def find_target_qty_column(df):
    cols = df.columns.tolist()
    numeric_cols = [
        c for c in cols
        if pd.api.types.is_numeric_dtype(df[c]) or _numeric_series(df, c).abs().sum() > 0
    ]
    if not numeric_cols:
        return None

    def score(col):
        n = _normalize_header(col)
        s = 0
        if "targetqty" in n or "targetquantity" in n: s += 120
        if "target" in n: s += 70
        if any(k in n for k in ["qty", "quantity", "unit", "uinit", "unt"]): s += 60
        if n.startswith("routewise"): s += 20
        if any(k in n for k in ["value", "vlue", "val", "amount", "taka", "tk", "code", "id"]): s -= 40
        return s

    ranked = sorted(numeric_cols, key=lambda c: (score(c), _numeric_series(df, c).abs().sum()), reverse=True)
    return ranked[0] if score(ranked[0]) > 0 else None

def find_preferred_text_column(cols, matcher, name_tokens):
    candidates = [c for c in cols if matcher(c)]
    if not candidates:
        return None

    def score(col):
        n = _normalize_header(col)
        s = 0
        if any(tok in n for tok in name_tokens): s += 100
        if n.endswith("name"): s += 40
        if any(k in n for k in ["code", "id", "no", "number"]): s -= 100
        return s

    return sorted(candidates, key=score, reverse=True)[0]

def top_target_table(df, group_col, target_col, label_col, value_label="Target Value", limit=5, pack_size_col=None):
    if not group_col or not target_col:
        return pd.DataFrame(columns=[label_col, value_label])
    work = df[[group_col, target_col]].copy()
    if pack_size_col and pack_size_col in df.columns:
        work[pack_size_col] = df[pack_size_col]
        work[label_col] = (
            work[group_col].astype(str).str.strip()
            + " - "
            + work[pack_size_col].astype(str).str.strip()
        )
        work[label_col] = (
            work[label_col]
            .str.replace(r"\s*-\s*nan$", "", regex=True)
            .str.replace(r"\s*-\s*$", "", regex=True)
            .str.strip()
        )
        key_col = label_col
    else:
        key_col = group_col
    work[target_col] = _numeric_series(work, target_col)
    work = work[work[key_col].notna()]
    if work.empty:
        return pd.DataFrame(columns=[label_col, value_label])
    out = (
        work.groupby(key_col, dropna=False)[target_col]
        .sum()
        .sort_values(ascending=False)
        .head(limit)
        .reset_index()
    )
    out.columns = [label_col, value_label]
    return out

def generate_area_summary_tables(combined_df_list):
    """Calculates Sheet 2 tables for each area workbook."""
    if not combined_df_list: return []
    main_df = combined_df_list[0][1]
    if main_df.empty: return []
    
    cols = main_df.columns.tolist()
    mkt_col = find_preferred_text_column(cols, is_market_column, ["marketname", "mktname", "thaname", "outletname"])
    prod_col = find_preferred_text_column(cols, is_product_column, ["productname", "prodname", "itemname", "skuname"])
    terri_col = find_preferred_text_column(cols, is_territory_column, ["territoryname", "terriname", "tsoname"])
    category_col = find_preferred_text_column(cols, is_category_column, ["categoryname", "catname", "segmentname"])
    pack_size_col = find_preferred_text_column(cols, is_pack_size_column, ["packsize", "size"])
    target_col = find_target_value_column(main_df)
    target_qty_col = find_target_qty_column(main_df)

    summaries = []

    if target_col:
        summaries.append(("table", "📍 Top 5 Markets by Value", top_target_table(main_df, mkt_col, target_col, "Market Name")))
        summaries.append(("table", "🗺️ Top 5 Territories by Target Value", top_target_table(main_df, terri_col, target_col, "Territory")))
        if category_col:
            summaries.append(("table", "🏷️ Category Wise Value Summary", top_target_table(main_df, category_col, target_col, "Category", limit=9999)))

    if target_qty_col:
        summaries.insert(2, ("table", "📦 Top 5 Products by Target Qty", top_target_table(main_df, prod_col, target_qty_col, "Product Name", "Target Qty", pack_size_col=pack_size_col)))
        
    return summaries

def truncate_sheet_name(name, max_len=31):
    s = str(name) if name is not None else "Sheet"
    return s[:max_len]

def detect_area_columns(sheets_dict):
    area_code_col = area_name_col = None
    for df in sheets_dict.values():
        for c in df.columns:
            if is_area_code_col(c): area_code_col = c
            if is_area_name_col(c): area_name_col = c
    return area_code_col, area_name_col

@st.cache_data(show_spinner=False)
def _read_area_file_bytes(name: str, blob: bytes) -> dict:
    ext = os.path.splitext(name.lower())[1]
    bio = BytesIO(blob)
    if ext == ".csv":
        return {"Sheet1": pd.read_csv(bio)}
    elif ext == ".xlsb":
        try:
            return pd.read_excel(bio, sheet_name=None, engine="pyxlsb")
        except Exception as e:
            raise RuntimeError(f"xlsb read failed (run: pip install pyxlsb): {e}") from e
    else:
        return pd.read_excel(bio, sheet_name=None, engine="openpyxl")

def merge_area_workbook_uploads(file_list):
    if not file_list: return {}
    files = list(file_list)
    if len(files) == 1:
        f = files[0]
        return _read_area_file_bytes(f.name, f.getvalue())
    merged = {}
    for f in files:
        stem = clean_filename(os.path.splitext(getattr(f,"name","") or "wb")[0])[:12] or "bk"
        wbmap = _read_area_file_bytes(f.name, f.getvalue())
        for sname, df in wbmap.items():
            raw_key = f"{stem}_{sname}"
            k = truncate_sheet_name(raw_key); base_k = k; n = 0
            while k in merged:
                n += 1; k = truncate_sheet_name(f"{base_k}_{n}")
            merged[k] = df
    return merged

def apply_area_workbook_format(wb):
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="thin"), bottom=Side(style="thin"))
    hfill = PatternFill("solid", fgColor="1e3a5f")
    tfill = PatternFill("solid", fgColor="1a4731")
    hfont = Font(bold=True, color="FFFFFF")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    num_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    txt_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for ws in wb.worksheets:
        ws.row_dimensions[1].height = 20
        hidden_cols = set()
        
        # Area Summary Sheet styling
        if ws.title == "Area Summary":
            for i, col in enumerate(ws.columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = 30
            title_rows = {
                cell.row
                for row in ws.iter_rows()
                for cell in row
                if cell.value and any(str(cell.value).startswith(k) for k in ["📍","📦","🗺️","🏷️"])
            }
            for row in ws.iter_rows():
                for cell in row:
                    is_header = cell.row - 1 in title_rows
                    if cell.row in title_rows:
                        cell.font = Font(bold=True, size=14, color="1e3a5f")
                    elif is_header:
                        cell.fill = hfill
                        cell.font = hfont
                        cell.alignment = h_align
                    if cell.value is not None:
                        cell.border = thin
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = num_align
                            if cell.number_format != "0%":
                                cell.number_format = "#,##0"
                        elif is_header:
                            cell.alignment = h_align
                        else:
                            cell.alignment = txt_align
            continue

        for i, cell in enumerate(ws[1], 1):
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = h_align
            if cell.value and "code" in _normalize_header(str(cell.value)):
                ws.column_dimensions[get_column_letter(i)].hidden = True
                hidden_cols.add(i)
        for row in ws.iter_rows(min_row=2):
            has_total = any(
                cell.value and "total" in str(cell.value).lower()
                for cell in row
            )
            for cell in row:
                if cell.value is not None:
                    cell.border = thin
                    if isinstance(cell.value, float):
                        cell.value = round(cell.value)
                    if has_total:
                        cell.fill = tfill
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = num_align
                    else:
                        cell.alignment = txt_align
        for i, col_cells in enumerate(ws.iter_cols(), 1):
            if i in hidden_cols:
                continue
            col_letter = get_column_letter(i)
            hdr_val = ws.cell(1, i).value
            max_len = len(str(hdr_val)) if hdr_val is not None else 4
            for cell in col_cells[1:]:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

def build_formatted_area_excel(combined_list):
    bio = BytesIO(); used_names = {}
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        for sheet, df in combined_list:
            base = truncate_sheet_name(sheet); name = base
            n = used_names.get(base, 0)
            if n:
                suffix = f"_{n}"
                name = truncate_sheet_name(base[:31-len(suffix)] + suffix)
            used_names[base] = n + 1
            df.to_excel(writer, sheet_name=name, index=False)
        
        summaries = generate_area_summary_tables(combined_list)
        if summaries:
            summary_sheet_name = "Area Summary"
            if summary_sheet_name not in writer.book.sheetnames:
                writer.book.create_sheet(summary_sheet_name)
            ws = writer.book[summary_sheet_name]
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 26
            for col in ["E","F","G","H","I","J","K","L","M","N","O","P"]:
                ws.column_dimensions[col].width = 13

            # Each section: 1 title row + 1 header row + up to 7 data rows + chart below
            # Generous vertical spacing to prevent any overlap
            layout = {
                "market":    {"row": 1,   "anchor": "E2",   "w": 14, "h": 9},
                "territory": {"row": 35,  "anchor": "E36",  "w": 14, "h": 9},
                "product":   {"row": 69,  "anchor": "E70",  "w": 14, "h": 9},
                "category":  {"row": 103, "anchor": "E104", "w": 14, "h": 9},
            }

            def _section_key(title):
                t = str(title).lower()
                if "market" in t: return "market"
                if "territor" in t: return "territory"
                if "product" in t: return "product"
                if "category" in t: return "category"
                return None

            for _kind, title, sdf in summaries:
                if sdf.empty:
                    continue
                key = _section_key(title)
                if key not in layout:
                    continue
                start_row = layout[key]["row"]
                ws.cell(row=start_row, column=1, value=title)

                for col_idx, col_name in enumerate(sdf.columns, 1):
                    ws.cell(row=start_row + 1, column=col_idx, value=col_name)
                for row_idx, row in enumerate(sdf.itertuples(index=False), start_row + 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                data_first = start_row + 2
                data_last = start_row + 1 + len(sdf)
                if len(sdf.columns) >= 2 and data_last >= data_first:
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.width = layout[key]["w"]
                    # Dynamic height based on data rows
                    chart.height = max(5, (len(sdf) + 2) * 0.6)
                    chart.title = str(title).replace("📍 ", "").replace("🗺️ ", "").replace("📦 ", "").replace("🏷️ ", "")
                    chart.y_axis.title = None
                    chart.x_axis.title = None
                    # Remove gridlines
                    chart.y_axis.majorGridlines = None
                    chart.x_axis.majorGridlines = None
                    # Remove chart border/line
                    gp = GraphicalProperties()
                    gp.ln = LineProperties(noFill=True)
                    chart.graphical_properties = gp
                    # No legend
                    chart.legend = None
                    # Data labels: show only value
                    chart.dataLabels = DataLabelList()
                    chart.dataLabels.showVal = True
                    chart.dataLabels.showCatName = True
                    chart.dataLabels.showSerName = False
                    chart.dataLabels.showPercent = False
                    chart.dataLabels.showLegendKey = False
                    chart.gapWidth = 110

                    data_ref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=data_last)
                    cat_ref = Reference(ws, min_col=1, min_row=data_first, max_row=data_last)
                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cat_ref)
                    ws.add_chart(chart, layout[key]["anchor"])

    bio.seek(0)
    wb = load_workbook(bio)
    apply_area_workbook_format(wb)
    out = BytesIO(); wb.save(out)
    return out.getvalue()

def unique_xlsx_path(folder, file_base):
    candidate = os.path.join(folder, f"{file_base}.xlsx")
    if not os.path.exists(candidate):
        return candidate
    n = 2
    while True:
        candidate = os.path.join(folder, f"{file_base} ({n}).xlsx")
        if not os.path.exists(candidate):
            return candidate
        n += 1

def run_area_split_from_sheets(sheets_dict, area_code_col, area_name_col, output_folder=None):
    area_codes = set()
    for df in sheets_dict.values():
        if area_code_col in df.columns:
            area_codes.update(df[area_code_col].dropna().astype(str).str.strip())
    outp = normalize_output_folder(output_folder)
    save_report = None; written = []; save_error = None
    if outp:
        try: os.makedirs(outp, exist_ok=True)
        except OSError as e: save_error = str(e)
    zip_buffer = BytesIO(); area_files = {}
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for code in sorted(area_codes, key=str):
            combined_df = []
            for sheet, df in sheets_dict.items():
                if area_code_col not in df.columns:
                    continue
                filtered = df[df[area_code_col].astype(str).str.strip() == code]
                if not filtered.empty:
                    combined_df.append((sheet, filtered))
            if not combined_df: continue
            area_name = combined_df[0][1][area_name_col].iloc[0]
            file_base = clean_filename(area_name)
            xlsx_bytes = build_formatted_area_excel(combined_df)
            zf.writestr(f"{file_base}.xlsx", xlsx_bytes)
            area_files[file_base] = combined_df
            if outp and save_error is None:
                out_path = unique_xlsx_path(outp, file_base)
                try:
                    with open(out_path, "wb") as f: f.write(xlsx_bytes)
                    written.append(out_path)
                except OSError as e: save_error = str(e)
    zip_buffer.seek(0)
    if outp and save_report is None:
        save_report = {"folder": outp, "written": written, "error": save_error}
    return zip_buffer.getvalue(), area_files, len(area_files), save_report

# =============================================================================
# STEP 6 — Outlook helpers
# =============================================================================
DEFAULT_OUTLOOK_FROM = "bijoy.laskor@squaregroup.com"
_APP_DIR = os.path.dirname(os.path.abspath(__file__))
OUTLOOK_LOG_PATH = os.path.join(_APP_DIR, "outlook_send_log.txt")

def outlook_environment_ok():
    """Checks if the system can send emails via Outlook or SMTP."""
    is_windows = (sys.platform == "win32")
    has_pywin32 = False
    if is_windows:
        try:
            import win32com.client; has_pywin32 = True
        except ImportError: pass
    
    # SMTP is always 'available' as it uses standard libs
    return True, {
        "is_windows": is_windows,
        "has_pywin32": has_pywin32,
        "can_outlook": is_windows and has_pywin32
    }

def append_outlook_log(lines):
    try:
        with open(OUTLOOK_LOG_PATH, "a", encoding="utf-8") as f:
            for l in lines: f.write(l + "\n")
    except OSError: pass

def outlook_sig(): return "Best regards,\nBijoy Laskor\nManagement Information System\nSquare Food & Beverage Ltd\n"
def outlook_body_all():
    return "Dear colleagues,\n\nPlease find your Area Wise Target in the attached file.\n\n" + outlook_sig()
def outlook_body_bhai(name):
    n = str(name).strip() if name else "Colleague"
    return f"Dear {n} Bhai,\n\nPlease find your Area Wise Target in the attached file.\n\n" + outlook_sig()

def format_subject(tpl, area_stem, name_str):
    t = (tpl or "").strip() or "Route Wise Target - {area}"
    a = str(area_stem).strip() if area_stem and str(area_stem).strip() else ""
    n = str(name_str).strip() if name_str and str(name_str).strip() else ""
    return t.replace("{area}", a).replace("{name}", n)

def merge_cc(row_cc, extra_cc):
    parts = []
    for src in [row_cc, extra_cc]:
        if src and str(src).strip():
            parts.extend([p.strip() for p in re.split(r"[;,]", str(src)) if p.strip()])
    seen = set(); out = []
    for p in parts:
        if p.lower() not in seen: seen.add(p.lower()); out.append(p)
    return "; ".join(out)

def find_area_payload(area_files, area_value):
    if area_value is None or (isinstance(area_value, float) and pd.isna(area_value)): return None, None
    raw = str(area_value).strip()
    if not raw: return None, None
    ck = clean_filename(raw)
    if ck in area_files: return ck, area_files[ck]
    for k in area_files:
        if k.lower() in (raw.lower(), ck.lower()): return k, area_files[k]
    return None, None

def area_payload_to_temp_file(payload, area_key):
    data = payload if isinstance(payload, list) else [("Sheet1", payload)]
    excel_bytes = build_formatted_area_excel(data)
    tdir = tempfile.mkdtemp()
    fname = clean_filename(area_key) + ".xlsx"
    path = os.path.join(tdir, fname)
    with open(path, "wb") as f:
        f.write(excel_bytes)
    return path, fname

def send_outlook_once(to_addr, cc_merged, mail_subject, mail_body, attachment_path):
    import pythoncom, win32com.client
    need_uninit = False
    try:
        pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED); need_uninit = True
    except pythoncom.com_error as e:
        if (int(getattr(e,"hresult",e.args[0])) & 0xFFFFFFFF) != 0x80010106: raise
    try:
        mail = win32com.client.Dispatch("Outlook.Application").CreateItem(0)
        mail.To = str(to_addr).strip()
        if cc_merged: mail.CC = cc_merged
        mail.Subject = str(mail_subject).strip() or "Route Wise Target"
        mail.Body = mail_body
        mail.Attachments.Add(attachment_path)
        mail.Send()
    finally:
        if need_uninit: pythoncom.CoUninitialize()

def send_smtp_once(to_addr, cc_merged, mail_subject, mail_body, attachment_path, cfg):
    """Sends email via SMTP (Office 365 / Gmail / etc). Works on Win & Mac."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    msg = MIMEMultipart()
    msg['From'] = cfg['user']
    msg['To'] = str(to_addr).strip()
    if cc_merged: msg['Cc'] = cc_merged
    msg['Subject'] = mail_subject

    msg.attach(MIMEText(mail_body, 'plain'))

    # Attachment
    if attachment_path and os.path.isfile(attachment_path):
        with open(attachment_path, "rb") as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(attachment_path)}"')
            msg.attach(part)

    # Combine recipients for SMTP send
    recipients = [str(to_addr).strip()]
    if cc_merged:
        recipients.extend([c.strip() for c in cc_merged.split(";") if c.strip()])

    with smtplib.SMTP(cfg['server'], cfg['port']) as server:
        server.starttls()
        server.login(cfg['user'], cfg['password'])
        server.sendmail(cfg['user'], recipients, msg.as_string())

def area_files_from_zip_bytes(zip_bytes):
    out = {}
    with zipfile.ZipFile(BytesIO(zip_bytes), "r") as zf:
        for info in zf.infolist():
            nm = info.filename
            if info.is_dir() or not nm.lower().endswith(".xlsx"): continue
            base = os.path.splitext(os.path.basename(nm))[0]
            key = clean_filename(base) or base
            with zf.open(info) as fp:
                sheets = pd.read_excel(fp, sheet_name=None, engine="openpyxl")
            out[key] = list(sheets.items())
    return out

def area_files_from_xlsx_uploads(file_list):
    out = {}
    for f in file_list:
        sheets = pd.read_excel(BytesIO(f.getvalue()), sheet_name=None, engine="openpyxl")
        key = clean_filename(os.path.splitext(f.name)[0]) or f.name
        out[key] = list(sheets.items())
    return out

def resolve_area_files_for_email(session_area, zip_file, xlsx_files, folder_path=None):
    merged = {}
    
    # 1. From Folder (NEW)
    if folder_path:
        norm = normalize_output_folder(folder_path)
        if norm and os.path.isdir(norm):
            for f in os.listdir(norm):
                if f.lower().endswith(".xlsx"):
                    fpath = os.path.join(norm, f)
                    try:
                        stem = os.path.splitext(f)[0]
                        # Read all sheets back into the list of tuples format
                        sh_map = pd.read_excel(fpath, sheet_name=None, engine="openpyxl")
                        merged[clean_filename(stem) or stem] = list(sh_map.items())
                    except: pass

    # 2. From ZIP
    if zip_file: merged.update(area_files_from_zip_bytes(zip_file.getvalue()))
    # 3. From XLSX Uploads
    if xlsx_files: merged.update(area_files_from_xlsx_uploads(xlsx_files))
    
    if merged: return merged
    if session_area: return dict(session_area)
    return {}

def find_area_payload_advanced(area_files, area_value, recipient_name):
    """Improved matching logic for names like 'khirul ASm Pabna Area'."""
    # 1. Direct match with Area value
    key, payload = find_area_payload(area_files, area_value)
    if payload: return key, payload
    
    # 2. Match Area value inside recipient name
    # (If the person has 'Pabna Area' in their name, find 'Pabna Area' file)
    rname = str(recipient_name).lower()
    for k, p in area_files.items():
        # Check if the filename (e.g. 'pabna_area') exists in some form in the name
        clean_k = str(k).replace("_", " ").lower()
        if clean_k in rname or str(k).lower() in rname:
            return k, p
            
    # 3. Last ditch: try to match any file key inside the area_value string
    raw_area = str(area_value).lower()
    for k, p in area_files.items():
        if str(k).lower() in raw_area:
            return k, p
            
    return None, None

def fuzzy_match_area_to_file(asm_area_name, area_file_keys, threshold=0.55):
    """Match an ASM's area name to area file keys using fuzzy matching.
    Returns (best_key, ratio) or (None, 0)."""
    if not asm_area_name or not area_file_keys:
        return None, 0
    raw = str(asm_area_name).strip().lower()
    raw_clean = raw.replace("_", " ")
    best_key, best_ratio = None, 0
    for k in area_file_keys:
        k_clean = str(k).replace("_", " ").lower().strip()
        # exact
        if raw_clean == k_clean:
            return k, 1.0
        # substring
        if raw_clean in k_clean or k_clean in raw_clean:
            return k, 0.95
        # fuzzy
        ratio = difflib.SequenceMatcher(None, raw_clean, k_clean).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_key = k
    if best_ratio >= threshold:
        return best_key, best_ratio
    return None, 0

def build_asm_area_matches(asm_df, asm_area_col, area_file_keys):
    """Build a list of (index, asm_area_value, matched_file_key, ratio) for each ASM row."""
    matches = []
    for idx, row in asm_df.iterrows():
        area_val = str(row[asm_area_col]).strip() if pd.notna(row[asm_area_col]) else ""
        mk, ratio = fuzzy_match_area_to_file(area_val, area_file_keys)
        matches.append({"idx": idx, "asm_area": area_val, "matched_key": mk, "ratio": ratio})
    return matches

def guess_col(columns, keywords, reject=()):
    for c in columns:
        n = _normalize_header(c)
        if any(r in n for r in reject): continue
        for kw in keywords:
            if n == kw: return c
    for c in columns:
        n = _normalize_header(c)
        if any(r in n for r in reject): continue
        if any(kw in n for kw in keywords): return c
    return None

# =============================================================================
# TABS
# =============================================================================
tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "1️⃣ Generate Template",
    "2️⃣ AVG & Contribution",
    "3️⃣ DB Target Merge",
    "4️⃣ Route Target Split",
    "5️⃣ Area-wise Split",
    "6️⃣ Auto Email",
    "7️⃣ Mid-Month SR Join",
])

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — Generate Template
# ─────────────────────────────────────────────────────────────────────────────
with tab1:
    step_header("📝", "Generate Template",
                "Cross-join Location × Product to produce master template used in all downstream steps")
    st.divider()
    st.markdown("### 📝 Template Generation Options")

    src1 = st.radio("Mode", ["🆕 Create New Template", "📥 Upload Existing Template"],
                    horizontal=True, key="t1_src")
    if src1 == "📥 Upload Existing Template":
        st.markdown('<div class="upload-section">', unsafe_allow_html=True)
        uf = st.file_uploader("📋 Template File (CSV / XLSX / XLSB)", type=["csv","xlsx","xlsb"], key="t1_upf")
        st.markdown('</div>', unsafe_allow_html=True)
        if uf:
            try:
                df_tpl = load_file(uf)
                ss_set("template", df_tpl); mark_step_done(1)
                st.success(f"✅ Template loaded — {len(df_tpl):,} rows × {len(df_tpl.columns)} cols")
                with st.expander("📊 Dashboard", expanded=True): mini_dashboard(df_tpl, key="t1u")
                c1,c2 = st.columns(2)
                with c1: download_csv(df_tpl, "template.csv", key="t1u_c")
                with c2: download_excel(df_tpl, "template.xlsx", key="t1u_x")
            except RuntimeError as e: st.error(str(e))
        else:
            st.info("Upload a template CSV/XLSX to continue.")
    else:
        st.markdown('<div class="upload-section">', unsafe_allow_html=True)
        c1,c2,c3 = st.columns(3)
        with c1: master_file = st.file_uploader("📦 Product Master (CSV / XLSX / XLSB)", type=["csv","xlsx","xlsb"], key="t1_mf")
        with c2: location_file = st.file_uploader("📍 Location Structure (CSV / XLSX / XLSB)", type=["csv","xlsx","xlsb"], key="t1_lf")
        with c3: db_map_file = st.file_uploader("🔗 DB Mapping (opt.) (CSV / XLSX / XLSB)", type=["csv","xlsx","xlsb"], key="t1_db")
        st.markdown('</div>', unsafe_allow_html=True)
        if master_file and location_file:
            try:
                df_master = load_file(master_file)
                df_location = load_file(location_file)
                df_db = load_file(db_map_file) if db_map_file else None
            except RuntimeError as e:
                st.error(str(e)); st.stop()
            with st.expander("📊 Source Dashboards", expanded=False):
                cl,cr = st.columns(2)
                with cl: mini_dashboard(df_master, "Product Master", "t1m")
                with cr: mini_dashboard(df_location, "Location Structure", "t1l")
            if df_db is not None:
                st.markdown("#### 🔗 DB Mapping Join")
                j1,j2 = st.columns(2)
                with j1: jloc = st.selectbox("Location join key", df_location.columns.tolist(), key="t1_jloc")
                with j2: jdb = st.selectbox("DB join key", df_db.columns.tolist(), key="t1_jdb")
                if st.button("Apply DB Mapping", key="t1_apply_db", type="secondary"):
                    df_location = df_location.merge(df_db, left_on=jloc, right_on=jdb, how="left")
                    st.success("DB mapping applied.")
            st.markdown("#### 📋 Column Selection")
            cl1,cl2 = st.columns(2)
            with cl1:
                loc_cols_sel = st.multiselect("Location columns", df_location.columns.tolist(),
                                              default=df_location.columns.tolist(), key="t1_locs")
            with cl2:
                prod_cols_sel = st.multiselect("Product columns", df_master.columns.tolist(),
                                               default=df_master.columns.tolist(), key="t1_prods")
            st.markdown("#### 🔍 Location Filter")
            filter_col = st.selectbox("Filter column", ["— No Filter —"]+df_location.columns.tolist(), key="t1_fc")
            df_loc_f = df_location.copy()
            if filter_col != "— No Filter —":
                if pd.api.types.is_numeric_dtype(df_location[filter_col]):
                    mn, mx = float(df_location[filter_col].min()), float(df_location[filter_col].max())
                    rng = st.slider("Range", mn, mx, (mn, mx), key="t1_rng")
                    df_loc_f = df_loc_f[(df_loc_f[filter_col]>=rng[0]) & (df_loc_f[filter_col]<=rng[1])]
                else:
                    uniq = sorted(df_location[filter_col].dropna().unique())
                    chosen = st.multiselect("Values", uniq, default=uniq, key="t1_vals")
                    if chosen: df_loc_f = df_loc_f[df_loc_f[filter_col].isin(chosen)]
                st.metric("Rows after filter", f"{len(df_loc_f):,}")
            if st.button("🚀 Generate Template", type="primary", key="t1_gen"):
                if not loc_cols_sel:
                    required_error("Select at least one Location column."); st.stop()
                if not prod_cols_sel:
                    required_error("Select at least one Product column."); st.stop()
                with st.spinner("Generating…"):
                    df_lu = df_loc_f[loc_cols_sel].drop_duplicates()
                    df_pu = df_master[prod_cols_sel].drop_duplicates()
                    t0 = time.time()
                    df_tpl = df_lu.merge(df_pu, how="cross")
                    elapsed = time.time() - t0
                ss_set("template", df_tpl); mark_step_done(1)
                st.success(f"✅ {len(df_tpl):,} rows in {elapsed:.2f}s")
                with st.expander("Preview", expanded=True):
                    st.dataframe(df_tpl.head(50), use_container_width=True)
                c1,c2 = st.columns(2)
                with c1: download_csv(df_tpl, "template.csv", key="t1_dlc")
                with c2: download_excel(df_tpl, "template.xlsx", key="t1_dlx")
        else:
            st.info("Upload Product Master and Location Structure to begin, or switch to Upload Existing Template mode above.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — AVG & Contribution
# ─────────────────────────────────────────────────────────────────────────────
with tab2:
    step_header("📊", "AVG & Contribution",
                "Pivot Sales → Avg (>0) → VLOOKUP Target → Contribution")

    # ── Individual Previous-Step Uploader ────────────────────────────────────
    tpl_df2 = prev_step_uploader(
        tab_key="t2",
        from_step=1,
        from_label="Step 1 — Generated Template",
        from_desc="The cross-joined Location × Product template produced in Tab 1.",
        sess_key="template",
        accepted_cols_hint="Route/Location · Product · DB Location",
    )

    st.divider()

    src2 = st.radio("Mode", ["🆕 Process Sales Data", "📥 Upload Existing Contribution"],
                    horizontal=True, key="t2_src")

    if src2 == "📥 Upload Existing Contribution":
        st.markdown('<div class="upload-section">', unsafe_allow_html=True)
        contrib_file2 = st.file_uploader("Contribution Result (CSV / XLSX / XLSB)", type=["csv","xlsx","xlsb"], key="t2_contrib")
        st.markdown('</div>', unsafe_allow_html=True)
        if contrib_file2:
            try:
                contrib_df2 = load_file(contrib_file2)
                ss_set("t2_result", contrib_df2); mark_step_done(2)
                st.success(f"✅ Loaded — {len(contrib_df2):,} rows")
                mini_dashboard(contrib_df2, "Existing Data", "t2u")
                c1,c2 = st.columns(2)
                with c1: download_csv(contrib_df2, "contribution_result.csv", key="t2u_c")
                with c2: download_excel(contrib_df2, "contribution_result.xlsx", key="t2u_x")
            except Exception as e: st.error(f"Load failed: {e}")
        else: st.info("Upload a file to continue.")

    else:
        # --- STEP 2.1: UPLOAD SALES DATA ---
        st.markdown("#### 1️⃣ Upload Sales Data")
        st.markdown('<div class="upload-section">', unsafe_allow_html=True)
        sales_file = st.file_uploader("Sales Data (12 Months Qty/Value) (CSV / XLSX / XLSB)", type=["csv","xlsx","xlsb"], key="t2_sf")
        st.markdown('</div>', unsafe_allow_html=True)

        if not sales_file:
            st.info("👆 Please upload your 12-month sales data file to begin.")
        else:
            with st.spinner("Reading file..."):
                raw_sales_df = _load_t2(sales_file.name, sales_file.getvalue())
            all_cols = raw_sales_df.columns.tolist()
            st.success(f"✅ {len(raw_sales_df):,} rows loaded.")

            # --- STEP 2.2: PIVOT & COLUMN MAPPING ---
            st.markdown("#### 2️⃣ Pivot & Column Mapping")
            st.info("Select the columns to group by (Pivot) and the month columns to sum/average.")
            
            k1,k2,k3 = st.columns(3)
            with k1: loc_col2 = st.selectbox("Route / Location Column", all_cols, index=_guess(all_cols,["route","territory","location"]), key="t2_loc")
            with k2: prod_col2 = st.selectbox("Product Column", all_cols, index=_guess(all_cols,["product","prod","item","sku","code"]), key="t2_prod")
            with k3: db_loc2 = st.selectbox("DB Location Column", all_cols, index=_guess(all_cols,["db loc","db location","distributor"]), key="t2_dbloc")

            month_sugg = get_column_suggestions(raw_sales_df, "month")
            month_cols2 = st.multiselect("📅 Select Month Columns (Sales Qty)",
                                          all_cols, default=month_sugg if month_sugg else None, key="t2_months")

            if st.button("🔄 Pivot & Calculate Average", type="primary", key="t2_pivot_btn"):
                if not month_cols2:
                    st.error("⚠️ Please select at least one month column.")
                else:
                    with st.spinner("Pivoting and calculating averages..."):
                        # Ensure numeric
                        for c in month_cols2:
                            raw_sales_df[c] = pd.to_numeric(raw_sales_df[c], errors="coerce").fillna(0)
                        
                        # 1. Pivot (Group By Sum)
                        pivoted = raw_sales_df.groupby([loc_col2, prod_col2, db_loc2], as_index=False)[month_cols2].sum()
                        
                        # 2. Calculate Average (>0 only)
                        # We use .where(gt(0)) so that 0s are excluded from mean()
                        mo_data = pivoted[month_cols2].where(pivoted[month_cols2] > 0)
                        pivoted["Route Avg Qty"] = mo_data.mean(axis=1).fillna(0)
                        
                        ss_set("t2_avg_data", pivoted)
                        ss_set("t2_mapping", {"loc": loc_col2, "prod": prod_col2, "db": db_loc2, "months": month_cols2})
                        st.success(f"✅ Pivoted to {len(pivoted):,} unique combinations.")
                        st.dataframe(pivoted.head(20), use_container_width=True)

            # --- STEP 2.3: TARGET MERGE & CONTRIBUTION ---
            avg_data = ss_get("t2_avg_data")
            if avg_data is not None:
                st.divider()
                st.markdown("### ✅ Pivot Data Ready")
                c1, c2 = st.columns([2, 1])
                with c1:
                    st.success(f"Pivoted data contains {len(avg_data):,} records.")
                with c2:
                    download_csv(avg_data, "pivoted_sales_avg.csv", label="📥 Download Pivoted CSV", key="t2_dl_pivot_res")
                
                st.markdown("#### 3️⃣ Target Format File (VLOOKUP)")
                st.markdown('<div class="upload-section">', unsafe_allow_html=True)
                fmt_file = st.file_uploader("Upload Target Format File (Template) (CSV / XLSX / XLSB)", type=["csv","xlsx","xlsb"], key="t2_fmt")
                st.markdown('</div>', unsafe_allow_html=True)
                
                use_sess = False
                if fmt_file is None and tpl_df2 is not None:
                    use_sess = st.checkbox("Use Step 1 template from session", value=True, key="t2_use_sess")
                
                if fmt_file or use_sess:
                    fmt_df = _load_fmt(fmt_file.name, fmt_file.getvalue()) if fmt_file else tpl_df2.copy().astype(str)
                    f_cols = fmt_df.columns.tolist()
                    
                    st.markdown("##### Match Columns in Target File")
                    c1,c2,c3 = st.columns(3)
                    with c1: f_loc = st.selectbox("Target - Route/Loc", f_cols, index=_guess(f_cols,["route","location"]), key="t2_floc")
                    with c2: f_prod = st.selectbox("Target - Product", f_cols, index=_guess(f_cols,["product","code"]), key="t2_fprod")
                    with c3: f_db = st.selectbox("Target - DB Location", f_cols, index=_guess(f_cols,["db location","distributor"]), key="t2_fdb")
                    
                    if st.button("🚀 Calculate Final Contribution", type="primary", key="t2_final_btn"):
                        try:
                            # Prepare AVG data for merge
                            mapping = ss_get("t2_mapping")
                            avg_subset = avg_data[[mapping["loc"], mapping["prod"], "Route Avg Qty"]].copy()
                            avg_subset.columns = ["_match_loc", "_match_prod", "Route Avg Qty"]
                            avg_subset["_match_loc"] = avg_subset["_match_loc"].astype(str).str.strip()
                            avg_subset["_match_prod"] = avg_subset["_match_prod"].astype(str).str.strip()
                            
                            # Prepare Target data
                            result = fmt_df.copy()
                            result["_t_loc"] = result[f_loc].astype(str).str.strip()
                            result["_t_prod"] = result[f_prod].astype(str).str.strip()
                            
                            # VLOOKUP (Left Merge)
                            result = result.merge(avg_subset, left_on=["_t_loc", "_t_prod"], right_on=["_match_loc", "_match_prod"], how="left")
                            result.drop(columns=["_t_loc", "_t_prod", "_match_loc", "_match_prod"], inplace=True)
                            result["Route Avg Qty"] = pd.to_numeric(result["Route Avg Qty"], errors="coerce").fillna(0)
                            
                            # Contribution Calculation
                            # Contribution = Route Avg / Sum(Avg for that DB + Product)
                            result["DB Sum Avg Qty"] = result.groupby([f_db, f_prod])["Route Avg Qty"].transform("sum")
                            result["Contribution"] = np.where(result["DB Sum Avg Qty"] > 0, 
                                                               result["Route Avg Qty"] / result["DB Sum Avg Qty"], 0.0)
                            
                            ss_set("t2_result", result)
                            mark_step_done(2)
                            st.success("✅ Contribution calculated successfully!")
                            
                            m1,m2,m3 = st.columns(3)
                            m1.metric("Total Rows", len(result))
                            m2.metric("Matched (>0 Avg)", (result["Route Avg Qty"] > 0).sum())
                            m3.metric("Unmatched", (result["Route Avg Qty"] == 0).sum())
                            
                            st.dataframe(result.head(30), use_container_width=True)
                            
                            d1, d2 = st.columns(2)
                            with d1: download_csv(result, "avg_contribution.csv", key="t2_dlc")
                            with d2: download_excel(result, "avg_contribution.xlsx", key="t2_dlx")
                            
                        except Exception as ex:
                            st.error(f"Error: {ex}")
                            st.exception(ex)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — DB Target Merge
# ─────────────────────────────────────────────────────────────────────────────
with tab3:
    step_header("🔗", "DB Target Merge (VLOOKUP)",
                "Match and bring DB-level targets into your Format + Contribution file")

    # ── Previous Step: Format + Contribution from Step 2 ──
    contrib_df3 = prev_step_uploader(
        tab_key="t3",
        from_step=2,
        from_label="Step 2 — Format + Contribution Result",
        from_desc="The AVG & Contribution output with Route Avg Qty, DB Sum, Contribution columns.",
        sess_key="t2_result",
        accepted_cols_hint="Route/Location · Product · DB Location · Contribution",
    )
    st.divider()

    st.markdown('**📁 Upload DB Target File** — contains DB-level targets to VLOOKUP')
    db_tgt_file3 = st.file_uploader("DB Target File (CSV / XLSX / XLSB)", type=["csv","xlsx","xlsb"], key="t3_dbf")

    if contrib_df3 is None:
        required_error("Upload or carry forward the Step 2 Format + Contribution result above.")
    elif db_tgt_file3 is None:
        required_error("Upload a DB Target file to merge.")
    else:
        try:
            db_df3 = load_file(db_tgt_file3)
        except RuntimeError as e:
            st.error(str(e)); st.stop()

        st.success(f"✅ DB Target: {len(db_df3):,} rows · Base File: {len(contrib_df3):,} rows")
        cont_cols3 = contrib_df3.columns.tolist()
        db_cols3 = db_df3.columns.tolist()

        st.markdown("#### 🗂️ VLOOKUP / Match Keys")
        st.caption("💡 We match based on DB Location and Product to bring selected columns.")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Format + Contribution Keys:**")
            fmt_db3 = st.selectbox("DB Location (Base)", cont_cols3,
                                   index=_guess(cont_cols3, ["db loc","db location","distributor"]), key="t3_fdb")
            fmt_prod3 = st.selectbox("Product (Base)", cont_cols3,
                                     index=_guess(cont_cols3, ["product","prod","item","sku","code"]), key="t3_fprod")
        with c2:
            st.markdown("**DB Target File Keys:**")
            tgt_db3 = st.selectbox("DB Location (Target)", db_cols3,
                                   index=_guess(db_cols3, ["db loc","db location","distributor"]), key="t3_tdb")
            tgt_prod3 = st.selectbox("Product (Target)", db_cols3,
                                     index=_guess(db_cols3, ["product","prod","item","sku","code"]), key="t3_tprod")

        st.markdown("#### 🎯 Select Columns to VLOOKUP")
        tgt_val_cols3 = st.multiselect("Pick columns to bring from DB Target file", db_cols3,
                                        default=[c for c in db_cols3 if any(k in c.lower() for k in ["target","tgt","qty","value","unit"])],
                                        key="t3_tgtcols")

        # --- KEY PREVIEW: show sample values so user can verify columns match ---
        with st.expander("🔍 Preview Match Keys (click to verify)", expanded=True):
            prev1, prev2 = st.columns(2)
            with prev1:
                st.markdown(f"**Base file → `{fmt_db3}` (sample):**")
                base_db_samples = contrib_df3[fmt_db3].astype(str).str.strip().drop_duplicates().head(10).tolist()
                st.code("\n".join(base_db_samples), language=None)
                st.markdown(f"**Base file → `{fmt_prod3}` (sample):**")
                base_prod_samples = contrib_df3[fmt_prod3].astype(str).str.strip().drop_duplicates().head(10).tolist()
                st.code("\n".join(base_prod_samples), language=None)
            with prev2:
                st.markdown(f"**DB Target → `{tgt_db3}` (sample):**")
                tgt_db_samples = db_df3[tgt_db3].astype(str).str.strip().drop_duplicates().head(10).tolist()
                st.code("\n".join(tgt_db_samples), language=None)
                st.markdown(f"**DB Target → `{tgt_prod3}` (sample):**")
                tgt_prod_samples = db_df3[tgt_prod3].astype(str).str.strip().drop_duplicates().head(10).tolist()
                st.code("\n".join(tgt_prod_samples), language=None)

        if st.button("🚀 Run DB Target VLOOKUP", type="primary", key="t3_merge"):
            if not tgt_val_cols3:
                required_error("Select at least one column to bring over."); st.stop()
            with st.spinner("Processing VLOOKUP…"):
                try:
                    # Robust matching - convert keys to stripped strings, lower for comparison
                    db_df_tmp = db_df3.copy()
                    db_df_tmp["_match_db"] = db_df_tmp[tgt_db3].astype(str).str.strip().str.lower()
                    db_df_tmp["_match_prod"] = db_df_tmp[tgt_prod3].astype(str).str.strip().str.lower()

                    # Aggregate to avoid row explosion if multiple entries exist for same DB+Prod
                    agg_map = {}
                    for col in tgt_val_cols3:
                        if pd.api.types.is_numeric_dtype(db_df_tmp[col]):
                            agg_map[col] = "sum"
                        else:
                            agg_map[col] = "first"
                    
                    db_agg3 = db_df_tmp.groupby(["_match_db", "_match_prod"], dropna=False).agg(agg_map).reset_index()

                    # Prepare Base file
                    base_tmp = contrib_df3.copy()
                    base_tmp["_base_db"] = base_tmp[fmt_db3].astype(str).str.strip().str.lower()
                    base_tmp["_base_prod"] = base_tmp[fmt_prod3].astype(str).str.strip().str.lower()

                    # Left Merge (VLOOKUP)
                    merged3 = base_tmp.merge(
                        db_agg3, 
                        left_on=["_base_db", "_base_prod"], 
                        right_on=["_match_db", "_match_prod"], 
                        how="left",
                        suffixes=("", "_target")
                    )

                    # Cleanup temporary match keys
                    merged3.drop(columns=["_base_db", "_base_prod", "_match_db", "_match_prod"], inplace=True)
                    
                    # Fill NaN for numeric columns
                    for c in tgt_val_cols3:
                        col_name = c if c in merged3.columns else (c + "_target" if c + "_target" in merged3.columns else None)
                        if col_name and pd.api.types.is_numeric_dtype(merged3[col_name]):
                            merged3[col_name] = pd.to_numeric(merged3[col_name], errors="coerce").fillna(0)

                    ss_set("t3_merged", merged3); mark_step_done(3)
                    
                    # --- DIAGNOSTIC: Match statistics ---
                    first_val_col = tgt_val_cols3[0]
                    check_col = first_val_col if first_val_col in merged3.columns else (first_val_col + "_target" if first_val_col + "_target" in merged3.columns else None)
                    if check_col:
                        matched_count = (pd.to_numeric(merged3[check_col], errors="coerce").fillna(0) != 0).sum()
                        unmatched_count = len(merged3) - matched_count
                    else:
                        matched_count = 0
                        unmatched_count = len(merged3)
                    
                    m1, m2, m3 = st.columns(3)
                    m1.metric("Total rows", f"{len(merged3):,}")
                    m2.metric("✅ Matched", f"{matched_count:,}")
                    m3.metric("❌ Unmatched (0)", f"{unmatched_count:,}")
                    
                    if unmatched_count > 0 and unmatched_count == len(merged3):
                        st.warning("⚠️ **All rows unmatched!** The key values may differ between files. Check the diagnostics below.")
                        with st.expander("🔎 Debugging: Why no matches?", expanded=True):
                            # Show what the keys look like after normalization
                            base_keys = set(
                                contrib_df3[fmt_db3].astype(str).str.strip().str.lower() + " | " +
                                contrib_df3[fmt_prod3].astype(str).str.strip().str.lower()
                            )
                            tgt_keys = set(
                                db_df3[tgt_db3].astype(str).str.strip().str.lower() + " | " +
                                db_df3[tgt_prod3].astype(str).str.strip().str.lower()
                            )
                            common = base_keys & tgt_keys
                            st.write(f"**Base unique key combos:** {len(base_keys):,}")
                            st.write(f"**Target unique key combos:** {len(tgt_keys):,}")
                            st.write(f"**Common (matched):** {len(common):,}")
                            
                            st.markdown("---")
                            d1, d2 = st.columns(2)
                            with d1:
                                st.markdown("**Sample Base keys (DB | Product):**")
                                for s in sorted(base_keys)[:15]:
                                    st.text(s)
                            with d2:
                                st.markdown("**Sample Target keys (DB | Product):**")
                                for s in sorted(tgt_keys)[:15]:
                                    st.text(s)
                    
                    st.success("✅ VLOOKUP complete — ready for Step 4 Route Target Split!")
                    st.dataframe(merged3.head(40), use_container_width=True)
                    
                    c1, c2 = st.columns(2)
                    with c1: download_csv(merged3, "db_target_vlookup_result.csv", key="t3_dlc")
                    with c2: download_excel(merged3, "db_target_vlookup_result.xlsx", key="t3_dlx")
                except Exception as ex:
                    st.error(f"Error during merge: {ex}")
                    st.exception(ex)

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — Route Target Split
# ─────────────────────────────────────────────────────────────────────────────
with tab4:
    step_header("🎯", "Route Target Split",
                "Upload any file with targets + contribution → split by VBA-exact logic")

    # ── Previous Step OR direct upload ──
    split_input_df = prev_step_uploader(
        tab_key="t4",
        from_step=3,
        from_label="Step 3 — DB Target Merged Result",
        from_desc="Output with contribution + DB targets merged.",
        sess_key="t3_merged",
        accepted_cols_hint="DB Location · Product · Route Segment · Product Segment · Contribution · Target columns",
    )
    st.divider()

    # Also allow direct file upload (independent mode)
    st.markdown("**📁 Or upload any file directly for route target split**")
    direct_upload4 = st.file_uploader("Direct File Upload (CSV / XLSX / XLSB)", type=["csv", "xlsx", "xlsb"], key="t4_direct")

    # Resolve which dataframe to use
    df4 = None
    if direct_upload4:
        try:
            df4 = load_file(direct_upload4)
            st.success(f"✅ Direct upload: {len(df4):,} rows × {len(df4.columns)} cols")
        except RuntimeError as e:
            st.error(str(e))
    elif split_input_df is not None:
        df4 = split_input_df

    if df4 is None:
        required_error("Upload a file above or carry forward Step 3 result.")
    else:
        cols4 = df4.columns.tolist()
        with st.expander("📊 Quick preview", expanded=False):
            st.dataframe(df4.head(10), use_container_width=True)

        st.markdown("#### 🗂️ Column Mapping")
        st.caption("💡 Columns are auto-suggested based on names. Adjust if needed.")
        c1, c2, c3 = st.columns(3)
        with c1:
            loc4 = st.selectbox("DB Location Column", cols4,
                                index=_guess(cols4, ["db loc","db location","location"]), key="t4_loc")
            prod4 = st.selectbox("Product Column", cols4,
                                 index=_guess(cols4, ["product","prod","item","sku","code"]), key="t4_prod")
        with c2:
            rseg4 = st.selectbox("Route Segment Column", cols4,
                                 index=_guess(cols4, ["route seg","route segment","segment"]), key="t4_rseg")
            pseg4 = st.selectbox("Product Segment Column", cols4,
                                 index=_guess(cols4, ["prod seg","product seg","product segment"]), key="t4_pseg")
        with c3:
            ctri4 = st.selectbox("Contribution Column", cols4,
                                 index=_guess(cols4, ["contribution","contri","contrib"]), key="t4_ctri")
            num4 = [
                c for c in cols4
                if pd.api.types.is_numeric_dtype(df4[c])
                or (len(df4) > 0 and vba_is_numeric(df4[c].iloc[0]))
            ]
            reject_tgt4 = ["avg", "contri", "contribution", "sum", "route wise", "total route"]
            tgt4_default = [
                c for c in num4
                if any(k in str(c).lower() for k in ["target","tgt","unit","value"])
                and not any(r in str(c).lower() for r in reject_tgt4)
            ]
            tgt4 = st.multiselect("Target Columns to Split", num4,
                                  default=tgt4_default if tgt4_default else [], key="t4_tgt")

        colA, colB = st.columns(2)
        with colA: rad4 = st.checkbox("Radhuni Strict Mode", value=True, key="t4_rad")
        with colB: ruc4 = st.checkbox("Ruchi Strict Mode", value=True, key="t4_ruc")

        if st.button("🚀 Run Route Target Split", type="primary", key="t4_run"):
            if not tgt4:
                required_error("Select at least one target column."); st.stop()
            with st.spinner("Running VBA-exact split…"):
                final_df4, t_src4, t_spl4, gaps4, ela4 = run_route_split(
                    df4, loc4, prod4, rseg4, pseg4, ctri4, tgt4, rad4, ruc4)
            ss_set("final_df", final_df4); mark_step_done(4)
            gap_qty4 = t_src4 - t_spl4
            m1, m2, m3 = st.columns(3)
            m1.metric("Total Source", f"{t_src4:,.0f}")
            m2.metric("Total Distributed", f"{t_spl4:,.0f}")
            m3.metric("Gap", f"{gap_qty4:,.0f}")
            st.caption(f"⏱️ {ela4:.2f}s")
            if gaps4:
                gdf4 = pd.DataFrame(list(gaps4.items()), columns=["Key", "Gap Qty"])
                gdf4 = gdf4[gdf4["Gap Qty"] > 0]
                if not gdf4.empty:
                    with st.expander(f"⚠️ Gap Details ({len(gdf4)} items)", expanded=True):
                        st.dataframe(gdf4, use_container_width=True)
            st.success(f"✅ Split complete — {len(final_df4):,} rows")
            st.dataframe(final_df4.head(50), use_container_width=True)
            c1, c2 = st.columns(2)
            with c1: download_csv(final_df4, "route_wise_target.csv", key="t4_dlc")
            with c2: download_excel(final_df4, "route_wise_target.xlsx", key="t4_dlx")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 — Area-wise Split
# ─────────────────────────────────────────────────────────────────────────────
with tab5:
    step_header("📍", "Area-wise Split",
                "Upload ANY file → split by Area Code → one formatted xlsx per area")

    st.markdown("""
    <div class="info-card">
      <strong>📂 Upload any file</strong> (xlsx / xlsb / csv) with Area Code + Area Name columns.<br>
      <strong>📊 Output:</strong> One formatted Excel per area — auto column-width · thin borders · total rows highlighted · code columns hidden
    </div>
    """, unsafe_allow_html=True)

    # ── DIRECT FILE UPLOAD — primary entry point ──
    st.markdown('<div class="upload-section">', unsafe_allow_html=True)
    area_workbooks = st.file_uploader(
        "📤 Upload file(s) to split by area — xlsx / xlsb / csv (Ctrl+click for multiple)",
        type=["xlsx", "xlsb", "csv"],
        accept_multiple_files=True,
        key="t5_area_wb",
    )
    st.markdown('</div>', unsafe_allow_html=True)

    # Also accept session data from Step 4 as fallback
    sheets5 = None
    if area_workbooks:
        try:
            with st.spinner("Reading files…"):
                sheets5 = merge_area_workbook_uploads(area_workbooks)
        except Exception as e:
            st.error(f"Could not read file(s): {e}")
    elif ss_get("final_df") is not None:
        if st.checkbox("📋 Or use Step 4 result from session", value=False, key="t5_use_sess"):
            sheets5 = {"Route Wise Target": ss_get("final_df")}
            st.info("Using Step 4 session data as source.")

    if sheets5:
        n_files = len(area_workbooks) if area_workbooks else 0
        n_sheets = len(sheets5)
        if n_files > 1:
            st.info(f"**{n_files}** files merged → **{n_sheets}** sheets total")
        else:
            st.info(f"**{n_sheets}** sheet(s) loaded")
        auto_code5, auto_name5 = detect_area_columns(sheets5)
        all_cols5 = list(dict.fromkeys(c for df in sheets5.values() for c in df.columns))
        st.markdown("#### ⚙️ Split Configuration")
        c1,c2 = st.columns(2)
        with c1:
            area_code_col5 = st.selectbox(
                "Area Code column (rows grouped by this)",
                all_cols5,
                index=all_cols5.index(auto_code5) if auto_code5 in all_cols5 else 0,
                key="t5_acode")
        with c2:
            area_name_col5 = st.selectbox(
                "Area Name column (used as output filename)",
                all_cols5,
                index=all_cols5.index(auto_name5) if auto_name5 in all_cols5 else min(1,len(all_cols5)-1),
                key="t5_aname")
        if auto_code5 and auto_name5:
            st.success(f"✅ Auto-detected — Code: `{auto_code5}` · Name: `{auto_name5}`")
        else:
            st.warning("Could not auto-detect area columns — please select manually.")
        all_codes5 = set()
        for df in sheets5.values():
            if area_code_col5 in df.columns:
                all_codes5.update(df[area_code_col5].dropna().astype(str).str.strip())
        st.metric("Unique Area Codes (= output files)", len(all_codes5))
        output_folder5 = st.text_input(
            "Output folder path (optional)",
            value=ss_get("t5_last_outf", ""),
            placeholder=r"E:\Route Wise Target\2026\April-26\Area Wise",
            key="t5_outf_manual",
            help="Paste any folder path. Leave blank for ZIP-only output."
        ).strip() or None
        if output_folder5:
            ss_set("t5_last_outf", output_folder5)
        if st.button("📤 Generate Area Files (ZIP)", type="primary", key="t5_gen"):
            if not area_code_col5 or not area_name_col5:
                required_error("Select both Area Code and Area Name columns."); st.stop()
            with st.spinner("Splitting and applying Excel formatting…"):
                zip_bytes5, area_files5, n5, save_rep5 = run_area_split_from_sheets(
                    sheets5, area_code_col5, area_name_col5,
                    output_folder=output_folder5)
            ss_set("area_zip_bytes", zip_bytes5)
            ss_set("area_files", area_files5)
            ss_set("area_zip_filename", "Area_Wise_Targets.zip")
            mark_step_done(5)
            st.success(f"✅ {n5} area workbook(s) created.")
            if save_rep5:
                if save_rep5.get("error"):
                    st.error(f"Folder save failed: {save_rep5['error']}")
                elif save_rep5.get("written"):
                    st.info(f"Saved {len(save_rep5['written'])} file(s) to `{save_rep5['folder']}`")
                    with st.expander("Saved paths"):
                        for p in save_rep5["written"]: st.code(p, language=None)
        if ss_get("area_zip_bytes"):
            st.download_button(
                "⬇️ Download All Area Files (ZIP)",
                ss_get("area_zip_bytes"),
                ss_get("area_zip_filename", "Area_Wise_Targets.zip"),
                "application/zip",
                key="t5_dlzip")
    else:
        required_error("Upload one or more workbook files (xlsx / xlsb / csv) to begin, or upload Step 4 result above.")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 6 — Auto Email via Outlook (Revamped)
# ─────────────────────────────────────────────────────────────────────────────
with tab6:
    step_header("📧", "Auto Email via Outlook",
                "Upload ASM mapping → attach area files → fuzzy match → send individually or all at once")

    env_ok, env_info = outlook_environment_ok()

    # ── 1. Sending Mode ──────────────────────────────────────────────────────
    st.markdown("""
    <div class="info-card">
      <strong>📨 Supported Modes:</strong><br>
      • <strong>Normal Mode</strong> — Outlook COM (Windows only, requires Outlook desktop app)<br>
      • <strong>Legacy / SMTP Mode</strong> — Works on Windows &amp; Mac (requires SMTP credentials)
    </div>
    """, unsafe_allow_html=True)

    methods = []
    if env_info.get("can_outlook"):
        methods.append("🖥️ Normal Mode (Outlook COM — Windows)")
    methods.append("🌐 Legacy Mode (SMTP — Windows & Mac)")

    email_method = st.radio("Sending Mode", methods, horizontal=True, key="t6_method")
    is_smtp_mode = "SMTP" in email_method or "Legacy" in email_method

    smtp_cfg = None
    if is_smtp_mode:
        smtp_cfg = {}
        with st.expander("🔐 SMTP Configuration", expanded=True):
            sc1, sc2 = st.columns(2)
            smtp_cfg['server'] = sc1.text_input("SMTP Server", value="smtp.office365.com", key="t6_sserv")
            smtp_cfg['port'] = int(sc2.number_input("Port", value=587, key="t6_sport"))
            smtp_cfg['user'] = sc1.text_input("Sender Email", value=DEFAULT_OUTLOOK_FROM, key="t6_suser")
            smtp_cfg['password'] = sc2.text_input("Password / App Password", type="password", key="t6_spass")
            st.caption("💡 For Office 365 use port 587. If 2FA is on, use an **App Password**.")
    else:
        st.caption(f"Sending via Outlook desktop app · Default sender: **{DEFAULT_OUTLOOK_FROM}**")

    st.divider()

    # ── 2. Area Files Input ──────────────────────────────────────────────────
    st.markdown("""
    <div class="prev-step-card">
      <div class="prev-step-card-title">📎 Step A — Area Split Files</div>
      <div class="prev-step-card-desc">
        Upload the area-split Excel files (ZIP from Tab 5, individual xlsx, or scan a folder).
        These will be attached to emails based on area matching.
      </div>
    </div>
    """, unsafe_allow_html=True)

    if ss_get("area_files"):
        n_af = len(ss_get("area_files"))
        st.markdown(
            f'<div class="session-banner">✅ {n_af} area file(s) loaded from Step 5 session'
            f'<span class="override-hint"> · Upload below to override</span></div>',
            unsafe_allow_html=True,
        )

    zu1, zu2 = st.columns(2)
    with zu1:
        zip_u6 = st.file_uploader("Area ZIP (from Step 5)", type=["zip"], key="t6_zip")
    with zu2:
        xlsx_u6 = st.file_uploader("Or individual area xlsx files", type=["xlsx"],
                                    accept_multiple_files=True, key="t6_xlsx")

    folder_path6 = st.text_input(
        "📂 Or scan local folder for Area files",
        placeholder=r"E:\Route Wise Target\2026\April-26\Area Wise",
        key="t6_folder",
        help="The app will look for .xlsx files in this folder."
    )

    area_files6 = resolve_area_files_for_email(
        ss_get("area_files"), zip_u6, xlsx_u6 if xlsx_u6 else None, folder_path=folder_path6
    )
    if area_files6:
        st.success(f"✅ {len(area_files6)} area file(s) ready for attachment")
        with st.expander("📂 Area file keys", expanded=False):
            st.write(sorted(area_files6.keys()))
    else:
        required_error("Upload area ZIP/xlsx or complete Step 5 in this session.")

    st.divider()

    # ── 3. ASM Mapping File ──────────────────────────────────────────────────
    st.markdown("""
    <div class="prev-step-card">
      <div class="prev-step-card-title">👤 Step B — ASM Mapping File</div>
      <div class="prev-step-card-desc">
        Upload a file with <strong>Area Name</strong>, <strong>ASM Name</strong>, and <strong>Mail ID</strong> columns.<br>
        <span style='color:#4a5568;font-size:0.72rem;'>The Area Name will be fuzzy-matched to area file names.</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    asm_file6 = st.file_uploader("ASM Mapping File (CSV / XLSX)", type=["csv", "xlsx", "xlsb"], key="t6_asm_file")
    df_asm6 = None
    if asm_file6:
        try:
            df_asm6 = load_file(asm_file6).dropna(how="all")
        except Exception as e:
            st.error(f"Could not read ASM file: {e}")

    if df_asm6 is not None and not df_asm6.empty:
        asm_cols = list(df_asm6.columns)
        g_asm_area = guess_col(asm_cols, ["areaname", "area"], ("code", "email", "mail", "cc"))
        g_asm_name = guess_col(asm_cols, ["asmname", "asm", "name", "fullname"], ("area", "email", "mail", "code", "cc"))
        g_asm_email = guess_col(asm_cols, ["email", "mail", "mailid", "emailid"])

        with st.expander("📋 ASM Column Mapping", expanded=True):
            mc1, mc2, mc3 = st.columns(3)
            with mc1:
                c_asm_area = st.selectbox("Area Name column", asm_cols,
                    index=asm_cols.index(g_asm_area) if g_asm_area in asm_cols else 0, key="t6_asm_area_col")
            with mc2:
                c_asm_name = st.selectbox("ASM Name column", asm_cols,
                    index=asm_cols.index(g_asm_name) if g_asm_name in asm_cols else 0, key="t6_asm_name_col")
            with mc3:
                c_asm_email = st.selectbox("Mail ID column", asm_cols,
                    index=asm_cols.index(g_asm_email) if g_asm_email in asm_cols else 0, key="t6_asm_email_col")

        st.success(f"✅ {len(df_asm6)} ASM recipient(s) loaded")

        # ── 4. Fuzzy Match Area Files ↔ ASM rows ────────────────────────────
        st.divider()
        st.markdown("#### 🔗 Area ↔ ASM Matching")

        if area_files6:
            file_keys = sorted(area_files6.keys())
            matches_raw = build_asm_area_matches(df_asm6, c_asm_area, file_keys)
            match_options = ["— None —"] + file_keys
            match_overrides = {}

            preview_data = []
            for m in matches_raw:
                row = df_asm6.loc[m["idx"]]
                preview_data.append({
                    "ASM Area": m["asm_area"],
                    "ASM Name": str(row[c_asm_name]) if pd.notna(row[c_asm_name]) else "",
                    "Email": str(row[c_asm_email]) if pd.notna(row[c_asm_email]) else "",
                    "Match": "✅" if m["matched_key"] else "❌",
                    "Matched File": m["matched_key"] or "—",
                    "Confidence": f"{m['ratio']*100:.0f}%" if m["matched_key"] else "—",
                })
            st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

            n_matched = sum(1 for m in matches_raw if m["matched_key"])
            n_total = len(matches_raw)
            st.caption(f"**{n_matched}/{n_total}** auto-matched. Use dropdowns below to override mismatches:")

            with st.expander("🔧 Override Matching (if needed)", expanded=(n_matched < n_total)):
                for m in matches_raw:
                    row = df_asm6.loc[m["idx"]]
                    lbl = f"{str(row[c_asm_name]) if pd.notna(row[c_asm_name]) else '?'} — {m['asm_area']}"
                    default_idx = match_options.index(m["matched_key"]) if m["matched_key"] in match_options else 0
                    sel = st.selectbox(lbl, match_options, index=default_idx, key=f"t6_override_{m['idx']}")
                    match_overrides[m["idx"]] = sel if sel != "— None —" else None

            final_matches = {}
            for m in matches_raw:
                ov = match_overrides.get(m["idx"])
                if ov is not None:
                    final_matches[m["idx"]] = ov
                elif m["matched_key"]:
                    final_matches[m["idx"]] = m["matched_key"]

            st.divider()

            # ── 5. CC / Subject / Body ───────────────────────────────────────
            st.markdown("#### ✏️ Email Composition")
            extra_cc6 = st.text_input("CC (semicolon-separated, added to all emails)", key="t6_xcc",
                placeholder="manager@squaregroup.com; boss@squaregroup.com")
            subj_base6 = st.text_input("Subject (area name auto-appended)", value="Route Wise Target",
                key="t6_subj_base", help="Final subject = your text + ' - ' + Area Name")
            body_text6 = st.text_area("Email Body",
                value="Dear Colleague,\n\nPlease find your Area Wise Target in the attached file.\n\n"
                      "Best regards,\nBijoy Laskor\nManagement Information System\nSquare Food & Beverage Ltd",
                height=180, key="t6_body")

            st.divider()

            # ── 6. Send Section ──────────────────────────────────────────────
            st.markdown("#### 📨 Send")
            send_mode6 = st.radio("Send mode", ["✉️ Send Individual", "📨 Send All"],
                                   horizontal=True, key="t6_mode")
            is_all6 = "All" in send_mode6

            def _resolve_attach_asm(idx):
                fk = final_matches.get(idx)
                if fk and fk in area_files6:
                    path, fname = area_payload_to_temp_file(area_files6[fk], fk)
                    return path, fk
                return None, None

            def _send_asm_row(idx, logs):
                row = df_asm6.loc[idx]
                to_a = str(row[c_asm_email]).strip() if pd.notna(row[c_asm_email]) else ""
                name = str(row[c_asm_name]).strip() if pd.notna(row[c_asm_name]) else ""
                area_nm = str(row[c_asm_area]).strip() if pd.notna(row[c_asm_area]) else ""
                ts = datetime.now().isoformat(timespec="seconds")
                if not to_a or "@" not in to_a:
                    logs.append(f"[{ts}] SKIP invalid email: {to_a}"); return False, "Invalid email"
                path, fk = _resolve_attach_asm(idx)
                if not path or not os.path.isfile(path):
                    logs.append(f"[{ts}] SKIP no attachment for: {name} ({area_nm})")
                    return False, f"No attachment matched for '{area_nm}'"
                area_label = area_nm or (fk.replace("_", " ") if fk else "")
                mail_subject = f"{subj_base6.strip()} - {area_label}".strip(" -")
                cc_m = extra_cc6.strip() if extra_cc6 else ""
                try:
                    if is_smtp_mode:
                        if not smtp_cfg or not smtp_cfg.get('password'):
                            raise ValueError("SMTP Password is required.")
                        send_smtp_once(to_a, cc_m, mail_subject, body_text6, path, smtp_cfg)
                    else:
                        send_outlook_once(to_a, cc_m, mail_subject, body_text6, path)
                    logs.append(f"[{ts}] ✔ Sent: {to_a} | {mail_subject}")
                    if path.startswith(tempfile.gettempdir()):
                        try: os.unlink(path)
                        except: pass
                    return True, None
                except Exception as ex:
                    logs.append(f"[{ts}] ❌ Error {to_a}: {ex}"); return False, str(ex)

            can_send = len(final_matches) > 0
            if not is_all6:
                sendable = [(idx, fk) for idx, fk in final_matches.items()]
                if not sendable:
                    st.warning("No ASM rows matched to area files. Fix matching above.")
                else:
                    def _asm_label(idx):
                        r = df_asm6.loc[idx]
                        nm = str(r[c_asm_name]) if pd.notna(r[c_asm_name]) else "?"
                        ar = str(r[c_asm_area]) if pd.notna(r[c_asm_area]) else ""
                        em = str(r[c_asm_email]) if pd.notna(r[c_asm_email]) else ""
                        return f"{nm} · {ar} → {final_matches.get(idx,'?')} · <{em}>"
                    pick_idx = st.selectbox("Select Recipient", [s[0] for s in sendable],
                                            format_func=_asm_label, key="t6_pick_asm")
                    if st.button("✉️ Send to Selected", type="primary", key="t6_send_one", disabled=not can_send):
                        logs6 = []
                        ok6, err6 = _send_asm_row(pick_idx, logs6)
                        append_outlook_log(logs6)
                        if ok6: st.success("✅ Sent — check Outlook Sent folder.")
                        else: st.error(err6 or "Failed.")
                        with st.expander("Log"): st.code("\n".join(logs6))
            else:
                st.info(f"Will send to **{len(final_matches)}** matched recipient(s).")
                if st.button("📨 Send All Mails", type="primary", key="t6_send_all", disabled=not can_send):
                    logs6 = []; ok_n = 0
                    prog6 = st.progress(0); total6 = len(final_matches)
                    with st.spinner("Sending…"):
                        for i, idx in enumerate(final_matches.keys()):
                            ok, _ = _send_asm_row(idx, logs6)
                            if ok: ok_n += 1
                            prog6.progress((i + 1) / max(total6, 1))
                    append_outlook_log(logs6)
                    st.success(f"✅ {ok_n}/{total6} emails sent.")
                    with st.expander("Log"): st.code("\n".join(logs6))
        else:
            st.info("👆 Upload area files in Step A above to enable matching.")
    else:
        if asm_file6:
            required_error("ASM mapping file is empty.")
        else:
            st.info("👆 Upload your ASM mapping file (Area Name · ASM Name · Mail ID) to begin.")


# ─────────────────────────────────────────────────────────────────────────────
# TAB 6 — Mid-Month SR Join: Dynamic Target Recalculation
# ─────────────────────────────────────────────────────────────────────────────
with tab7:
    step_header("🔄", "Mid-Month SR Join — Dynamic Target Adjustment",
                "Upload market targets → contributions auto-calculate → configure new SR → get pro-rated day-wise targets → download")

    st.markdown("""
    <div class="info-card">
      <strong>📐 Logic:</strong><br>
      Completed base = (join_day ÷ total_days) × Total target | Remaining base = rest<br>
      Existing market = (original% × completed_base) + (updated% × remaining_base)<br>
      New SR = new% × remaining_base ← pro-rated for remaining days only
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="upload-section">', unsafe_allow_html=True)
    st.markdown("**📁 Market-wise Target File**")
    st.caption("Required columns: **Market** (name/code), **Target** (monthly target value). "
               "Contribution % will be auto-calculated from the target values.")
    mkt_file7 = st.file_uploader("Market Target File (CSV / XLSX / XLSB)", type=["csv","xlsx","xlsb"], key="t7_mf")
    st.markdown('</div>', unsafe_allow_html=True)

    if not mkt_file7:
        required_error("Upload a market target file to begin. Format: Market | Target (tk value)")
        st.stop()

    df_mkt7 = load_file(mkt_file7)
    if df_mkt7 is None or df_mkt7.empty:
        st.error("Could not read file or file is empty."); st.stop()

    st.markdown("#### 🗂️ Column Mapping")
    mcols7 = df_mkt7.columns.tolist()
    cm1, cm2 = st.columns(2)
    with cm1:
        mkt_col7 = st.selectbox("Market Name / Code column", mcols7,
                                index=_guess(mcols7, ["market","name","route","location","area","sr"]),
                                key="t7_mktc")
    with cm2:
        tgt_col7 = st.selectbox("Target Value column", mcols7,
                                index=_guess(mcols7, ["target","value","qty","amount","tgt","tk"]),
                                key="t7_tgtc")

    df7 = df_mkt7[[mkt_col7, tgt_col7]].copy()
    df7.columns = ["Market", "Target"]
    df7["Target"] = pd.to_numeric(df7["Target"], errors="coerce").fillna(0)
    df7 = df7[df7["Target"] > 0].reset_index(drop=True)

    if df7.empty:
        required_error("No valid numeric target values found in the selected column."); st.stop()

    total_target_auto7 = df7["Target"].sum()
    df7["Contribution %"] = (df7["Target"] / total_target_auto7 * 100).round(4)

    st.success(f"✅ {len(df7)} markets loaded · Auto-calculated Total Target = **{total_target_auto7:,.2f}**")
    with st.expander("📊 Current Contributions (auto-calculated)", expanded=True):
        disp7 = df7.copy()
        disp7["Contribution %"] = disp7["Contribution %"].apply(lambda x: f"{x:.2f}%")
        st.dataframe(disp7, use_container_width=True, hide_index=True)

    st.markdown("#### ⚙️ Month Configuration")
    cfg1, cfg2, cfg3 = st.columns(3)
    with cfg1:
        total_days7 = st.number_input("Total days in month", min_value=1, max_value=31, value=30, step=1, key="t7_td")
    with cfg2:
        join_day7 = st.number_input("New SR joined on day", min_value=1, max_value=int(total_days7)-1,
                                        value=min(12, int(total_days7)-1), step=1, key="t7_jd",
                                        help="Days 1 to this day = completed. From next day = remaining.")
    with cfg3:
        override_tot7 = st.number_input("Override total target (0 = use file sum)", min_value=0.0,
                                        value=0.0, step=1.0, key="t7_ot")

    total_target7 = float(override_tot7) if override_tot7 > 0 else total_target_auto7
    comp_days7 = int(join_day7)
    rem_days7 = int(total_days7) - comp_days7
    comp_base7 = (comp_days7 / total_days7) * total_target7
    rem_base7 = (rem_days7 / total_days7) * total_target7

    mm1, mm2, mm3, mm4, mm5 = st.columns(5)
    mm1.metric("Total Target", f"{total_target7:,.1f}")
    mm2.metric("Total Days", f"{int(total_days7)}")
    mm3.metric("Completed Days", f"{comp_days7}")
    mm4.metric("Remaining Days", f"{rem_days7}")
    mm5.metric("Remaining Base (tk)", f"{rem_base7:,.1f}")

    st.markdown(f"#### ➕ New SR(s) — joining from Day {comp_days7 + 1}")
    n_new7 = st.number_input("How many new SRs joining?", min_value=0, max_value=10,
                              value=1, step=1, key="t7_nsr")
    new_srs7 = []
    if n_new7 > 0:
        for i in range(int(n_new7)):
            nc1, nc2 = st.columns(2)
            with nc1:
                sr_name7 = st.text_input(f"New SR {i+1} — Market name",
                                         value=f"New-SR-{i+1}", key=f"t7_srn_{i}")
            with nc2:
                sr_contrib7 = st.number_input(
                    f"New SR {i+1} — contribution % (based on full {int(total_days7)}-day month)",
                    min_value=0.0, max_value=100.0, value=20.0, step=0.1, key=f"t7_src_{i}",
                    help="This % represents the SR's weight for a full month.")
            new_srs7.append({"Market": sr_name7, "contrib_pct": sr_contrib7})

    new_sr_total_pct7 = sum(sr["contrib_pct"] for sr in new_srs7)
    st.markdown(f"#### 🔄 Updated Contributions — Day {comp_days7 + 1} to Day {int(total_days7)}")
    st.caption(f"New SR(s) take **{new_sr_total_pct7:.1f}%** → remaining for existing markets: **{100 - new_sr_total_pct7:.1f}%**")

    remaining_for_existing7 = 100.0 - new_sr_total_pct7
    if remaining_for_existing7 < 0:
        st.error(f"New SR contributions total {new_sr_total_pct7:.1f}% which exceeds 100%! Reduce them.")
        st.stop()

    exist_contrib_sum7 = df7["Contribution %"].sum()
    suggested7 = (df7["Contribution %"] / exist_contrib_sum7 * remaining_for_existing7
                  ).round(2) if exist_contrib_sum7 > 0 else pd.Series(
                  [remaining_for_existing7 / len(df7)] * len(df7))

    updated_contribs7 = []
    n_per_row = min(len(df7), 4)
    upd_cols7 = st.columns(n_per_row)
    for i, row in df7.iterrows():
        col_idx = i % n_per_row
        with upd_cols7[col_idx]:
            val7 = st.number_input(
                f"{row['Market']} updated %",
                min_value=0.0, max_value=100.0,
                value=float(round(suggested7.iloc[i], 2)),
                step=0.1, key=f"t7_uc_{i}",
                help=f"Original: {row['Contribution %']:.2f}% → suggested: {suggested7.iloc[i]:.2f}%")
            updated_contribs7.append(val7)

    exist_upd_sum7 = sum(updated_contribs7)
    grand_total_pct7 = exist_upd_sum7 + new_sr_total_pct7
    vok = abs(grand_total_pct7 - 100) < 0.5
    vc1, vc2, vc3 = st.columns(3)
    vc1.metric("Existing markets total %", f"{exist_upd_sum7:.1f}%")
    vc2.metric("New SR(s) total %", f"{new_sr_total_pct7:.1f}%")
    vc3.metric("Grand total %", f"{grand_total_pct7:.1f}%",
               delta="✓ OK" if vok else f"{grand_total_pct7-100:+.2f}%",
               delta_color="normal" if vok else "inverse")
    if not vok:
        st.warning(f"⚠️ Updated contributions sum to {grand_total_pct7:.1f}% — adjust to reach 100%.")

    st.divider()
    if st.button("⚡ Calculate Recalculated Targets", type="primary", key="t7_calc"):
        if not vok:
            required_error("Adjust contributions so grand total = 100%."); st.stop()
        results7 = []
        for i, row in df7.iterrows():
            orig_p = row["Contribution %"] / 100
            upd_p = updated_contribs7[i] / 100
            comp_t = round(orig_p * comp_base7, 4)
            rem_t = round(upd_p * rem_base7, 4)
            final_t = round(comp_t + rem_t, 4)
            results7.append({
                "Market": row["Market"],
                "Original Target": round(row["Target"], 2),
                "Original %": round(row["Contribution %"], 2),
                "Updated % (remaining days)": round(updated_contribs7[i], 2),
                f"Completed (Day 1–{comp_days7})": round(comp_t, 2),
                f"Remaining (Day {comp_days7+1}–{int(total_days7)})": round(rem_t, 2),
                "Final Target": round(final_t, 2),
                "Joining Day": 1,
            })
        for sr in new_srs7:
            sr_p = sr["contrib_pct"] / 100
            rem_t = round(sr_p * rem_base7, 4)
            results7.append({
                "Market": sr["Market"],
                "Original Target": 0,
                "Original %": 0,
                "Updated % (remaining days)": round(sr["contrib_pct"], 2),
                f"Completed (Day 1–{comp_days7})": 0,
                f"Remaining (Day {comp_days7+1}–{int(total_days7)})": round(rem_t, 2),
                "Final Target": round(rem_t, 2),
                "Joining Day": comp_days7 + 1,
            })
        result_df7 = pd.DataFrame(results7)
        total_final7 = result_df7["Final Target"].sum()
        sm1, sm2, sm3 = st.columns(3)
        sm1.metric("Original Total", f"{total_target7:,.2f}")
        sm2.metric("Recalculated Total", f"{total_final7:,.2f}")
        sm3.metric("Difference", f"{total_final7 - total_target7:+.4f}",
                   delta_color="normal" if abs(total_final7 - total_target7) < 0.01 else "inverse")
        st.markdown("#### 📊 Recalculated Target Table")
        disp_cols7 = [
            "Market", "Original Target", "Original %", "Updated % (remaining days)",
            f"Completed (Day 1–{comp_days7})",
            f"Remaining (Day {comp_days7+1}–{int(total_days7)})",
            "Final Target",
        ]
        st.dataframe(result_df7[disp_cols7], use_container_width=True, hide_index=True)
        st.markdown("#### 📅 Day-wise Target Breakdown")
        all_markets7 = list(df7["Market"]) + [sr["Market"] for sr in new_srs7]
        daily_rows7 = []
        for day in range(1, int(total_days7) + 1):
            dr = {"Day": day, "Period": f"Day 1–{comp_days7}" if day <= comp_days7 else f"Day {comp_days7+1}–{int(total_days7)}"}
            day_total = 0.0
            for i, mkt in enumerate(df7["Market"]):
                p = (df7.iloc[i]["Contribution %"]/100 if day <= comp_days7 else updated_contribs7[i]/100)
                daily = round(p * total_target7 / total_days7, 4)
                dr[mkt] = daily; day_total += daily
            for sr in new_srs7:
                if day <= comp_days7:
                    dr[sr["Market"]] = 0.0
                else:
                    daily = round((sr["contrib_pct"]/100) * total_target7 / total_days7, 4)
                    dr[sr["Market"]] = daily; day_total += daily
            dr["Daily Total"] = round(day_total, 4)
            daily_rows7.append(dr)
        daily_df7 = pd.DataFrame(daily_rows7)
        cum_df7 = daily_df7.copy()
        for col in all_markets7 + ["Daily Total"]:
            if col in cum_df7.columns:
                cum_df7[col] = cum_df7[col].cumsum().round(4)
        cum_df7 = cum_df7.rename(columns={"Day": "Day (Cumulative)"})
        tab_d, tab_c = st.tabs(["📅 Daily", "📈 Cumulative"])
        with tab_d:
            st.dataframe(daily_df7, use_container_width=True, hide_index=True,
                         column_config={"Period": st.column_config.TextColumn(width="medium")})
        with tab_c:
            st.dataframe(cum_df7, use_container_width=True, hide_index=True)
        ss_set("t7_result", result_df7)
        ss_set("t7_daily", daily_df7)
        ss_set("t7_cum", cum_df7)
        mark_step_done(7)
        st.markdown("#### ⬇️ Download")
        buf7 = BytesIO()
        with pd.ExcelWriter(buf7, engine="openpyxl") as wr7:
            result_df7[disp_cols7].to_excel(wr7, sheet_name="Recalculated Targets", index=False)
            daily_df7.to_excel(wr7, sheet_name="Daily Breakdown", index=False)
            cum_df7.to_excel(wr7, sheet_name="Cumulative", index=False)
        buf7.seek(0)
        wb7 = load_workbook(buf7)
        apply_area_workbook_format(wb7)
        out7 = BytesIO(); wb7.save(out7)
        dl1, dl2, dl3 = st.columns(3)
        with dl1:
            st.download_button(
                "📊 Download Excel (3 sheets)",
                out7.getvalue(),
                "mid_month_sr_adjustment.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="t7_dlx")
        with dl2:
            download_csv(result_df7[disp_cols7], "mid_month_sr_adjustment.csv", key="t7_dlc")
        with dl3:
            download_csv(daily_df7, "mid_month_daily_breakdown.csv", key="t7_dlc2")

# =============================================================================
# SIDEBAR
# =============================================================================
st.sidebar.markdown("""
<div class="hero-banner" style="flex-direction:column;align-items:flex-start;gap:0.4rem;padding:1rem 1.2rem;">
  🎯
  <div>
    <h1 class="hero-title" style="font-size:1.1rem;">Route Wise Target</h1>
    <p class="hero-sub">Automation Suite — v2.2</p>
  </div>
</div>
""", unsafe_allow_html=True)
st.sidebar.divider()
st.sidebar.markdown("### 📋 Workflow Progress")
steps_info = [
    (1, "📝", "Generate Template", "template"),
    (2, "📊", "AVG & Contribution", "t2_result"),
    (3, "🔗", "DB Target Merge", "t3_merged"),
    (4, "🎯", "Route Target Split", "final_df"),
    (5, "📍", "Area-wise Split", "area_zip_bytes"),
    (6, "📧", "Auto Email", None),
    (7, "🔄", "Mid-Month SR Join", "t7_result"),
]
for sn, icon, label, sk in steps_info:
    is_done = step_done(sn) or (sk and ss_get(sk) is not None)
    badge = "✅" if is_done else "⏳"
    cls = "done" if is_done else ""
    st.sidebar.markdown(
        f'<div class="sidebar-step {cls}">{badge} {sn}. {icon} {label}</div>',
        unsafe_allow_html=True)
st.sidebar.divider()
st.sidebar.markdown("### 📦 Session Data")
sess_items = [
    ("template", "Step 1 Template"),
    ("t2_result", "Step 2 Contribution"),
    ("t3_merged", "Step 3 Merged"),
    ("final_df", "Step 4 Split"),
    ("area_zip_bytes", "Step 5 ZIP"),
    ("t7_result", "Step 7 SR Adjust"),
]
any_data = False
for k, label in sess_items:
    v = ss_get(k)
    if v is not None:
        any_data = True
        rows = f"{len(v):,} rows" if isinstance(v, pd.DataFrame) else "ready"
        st.sidebar.markdown(f'<div class="chip green">✅ {label}: {rows}</div>',
                            unsafe_allow_html=True)
if not any_data:
    st.sidebar.caption("No session data yet — start from Tab 1 or upload in any tab.")
st.sidebar.divider()
st.sidebar.markdown("### ⚡ Actions")
if st.sidebar.button("🗑️ Clear All Session Data"):
    st.cache_data.clear(); st.session_state.clear()
    st.sidebar.success("Cleared!"); st.rerun()
if st.sidebar.button("🔄 Clear File Cache Only"):
    st.cache_data.clear(); st.sidebar.success("Cache cleared!"); st.rerun()

guide = """# Route Wise Target Automation — User Guide v2.2

## Individual Step Uploaders (NEW in v2.2)
Every tab now has a dedicated "Previous Step Input" card at the top.
Each card:
  • Shows exactly WHICH step's output it expects
  • Shows a green session banner if data is already loaded from a previous run
  • Accepts CSV or Excel upload to override session data
  • Gives a column hint so you know what the file must contain

This means you can jump to ANY tab independently — just upload the correct
previous step result file.

## Step 1: Generate Template
Upload Product Master + Location Structure.
Select columns, filter, Generate → download CSV/xlsx.

## Step 2: AVG & Contribution
  Previous Step Card: Upload Step 1 Template (CSV/xlsx)
  Upload sales data → set filters → map columns → Calculate Average.
  Upload Format/Template (or check 'use Step 1 session') → Calculate Contribution.

## Step 3: DB Target Merge
  Previous Step Card: Upload Step 2 Contribution Result (CSV/xlsx)
  Upload DB target file → Merge → CSV download.

## Step 4: Route Target Split
  Previous Step Card: Upload Step 3 Merged Result (CSV/xlsx)
  Map columns, set modes → Run Split → CSV download.

## Step 5: Area-wise Split
  Previous Step Card: Upload Step 4 Split Result (CSV/xlsx)
  Upload workbook(s) to split (or use Step 4 session directly via checkbox).
  Generate ZIP of formatted per-area xlsx files.

## Step 6: Auto Email (Windows + Outlook only)
  Previous Step Card: Upload Step 5 Area ZIP or individual xlsx files.
  Upload recipient list → Send All or Individual.

## Step 7: Mid-Month SR Join
  Standalone — upload market target file, configure SR, download results.

## Tips
- The "Previous Step Input" card at the top of each tab is the MAIN entry point
- Session data auto-carries forward; upload to override
- ALL column selections default to ALL columns — remove unwanted ones
- xlsb requires: pip install pyxlsb
- Clear cache in sidebar if you encounter stale data
"""
st.sidebar.download_button("📥 User Guide", guide, "user_guide.txt", "text/plain", key="sb_guide")
st.sidebar.divider()
st.sidebar.caption("Square Food & Beverage Ltd — MIS")
